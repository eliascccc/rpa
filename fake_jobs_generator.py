from __future__ import annotations
import random, os, threading, datetime, time, json, uuid, platform
from pathlib import Path
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from openpyxl import load_workbook #type: ignore

# to create fake email jobs
class FakeEmailjobsGenerator:
    #written by AI
    BASE_DIR = Path(__file__).resolve().parent
    PIPELINE_DIR = BASE_DIR / "personal_inbox"
    INBOX_DIR = PIPELINE_DIR / "inbox"
    PROCESSING_DIR =  PIPELINE_DIR / "processing"
    ATTACHMENTS_DIR = PIPELINE_DIR / "generator_attachments"

    for folder in [PIPELINE_DIR, INBOX_DIR, PROCESSING_DIR, ATTACHMENTS_DIR]:
        folder.mkdir(exist_ok=True)


    def __init__(self) -> None:
        self.main()


    def create_example_attachment_files(self) -> None:
        """Creates a few simple test files if they do not already exist."""
        txt_path = self.ATTACHMENTS_DIR / "job1_request.txt"
        if not txt_path.exists():
            txt_path.write_text(
                "SKU=100245\nOLD_MATERIAL=MAT-OLD-778\nNEW_MATERIAL=MAT-NEW-991\n",
                encoding="utf-8",
            )

        csv_path = self.ATTACHMENTS_DIR / "job2_request.csv"
        if not csv_path.exists():
            csv_path.write_text(
                "invoice_id,action\nINV-2026-1001,close\n",
                encoding="utf-8",
            )


    def build_email_message(self,
        *,
        from_name: str,
        from_email: str,
        to_email: str,
        subject: str,
        body: str,
        attachment_paths: list[Path] | None = None,
    ) -> EmailMessage:
        msg = EmailMessage()
        msg["From"] = f"{from_name} <{from_email}>"
        msg["To"] = to_email
        msg["Subject"] = subject
        msg["Date"] = formatdate(localtime=True)
        msg["Message-ID"] = make_msgid()
        msg.set_content(body)

        for path in attachment_paths or []:
            data = path.read_bytes()
            # Simple generic attachment type is enough for testing
            msg.add_attachment(
                data,
                maintype="application",
                subtype="octet-stream",
                filename=path.name,
            )

        return msg


    def write_eml_to_inbox(self,msg: EmailMessage, prefix: str = "mail") -> Path:
        """Atomic write into inbox to reduce risk of partial reads."""
        unique_id = uuid.uuid4().hex[:12]
        final_path = self.INBOX_DIR / f"{prefix}_{unique_id}.eml"
        temp_path = self.INBOX_DIR / f".tmp_{prefix}_{unique_id}.eml"

        with open(temp_path, "wb") as f:
            f.write(msg.as_bytes())

        temp_path.replace(final_path)
        return final_path


    def create_ping_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="PING",
            body="ping",
        )
        return self.write_eml_to_inbox(msg, prefix="ping")


    def create_job1_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Alice Wonderland",
            from_email="alice@example.com",
            to_email="robot@company.local",
            subject="Please run job1",
            body=(
                "I have no idea what job1 is though...\n"
                "Best regards,\n"
                "Alice\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job1_request.txt"],
        )
        return self.write_eml_to_inbox(msg, prefix="job1")

    def create_job1_b_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="Job1",
            body=(
                "Hello,\n\n"
                "Please run job1\n\n"
                "order_number: 100245\n"
                "order_qty: 12000\n"
                "material_available: 11031\n\n"
                "Best regards,\n"
                "Bob\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job1_request.txt"],
        )
        return self.write_eml_to_inbox(msg, prefix="job1")


    def create_job2_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Bob Tester",
            from_email="bob@test.com",
            to_email="robot@company.local",
            subject="Job2 request",
            body=(
                "Hello,\n\n"
                "Please run job2 using attached file.\n\n"
                "Regards,\n"
                "Bob\n"
            ),
            attachment_paths=[self.ATTACHMENTS_DIR / "job2_request.csv"],
        )
        return self.write_eml_to_inbox(msg, prefix="job2")


    def create_unknown_job_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Charlie Strange",
            from_email="charlie@example.com",
            to_email="robot@company.local",
            subject="Do some weird magic",
            body=(
                "Hello,\n\n"
                "Please do that strange thing the robot probably cannot classify.\n\n"
                "Regards,\n"
                "Charlie\n"
            ),
        )
        return self.write_eml_to_inbox(msg, prefix="unknown")


    def create_blocked_sender_mail(self) -> Path:
        msg = self.build_email_message(
            from_name="Mallory Intruder",
            from_email="mallory@evil.com",
            to_email="robot@company.local",
            subject="Please run job1",
            body=(
                "Hello,\n\n"
                "I would like the robot to run job1.\n\n"
                "Regards,\n"
                "Mallory\n"
            ),
        )
        return self.write_eml_to_inbox(msg, prefix="blocked")


    def create_random_mail(self) -> Path:
        creators = [
            self.create_ping_mail,
            self.create_job1_mail,
            self.create_job1_b_mail,  # valid
            self.create_unknown_job_mail,
            self.create_blocked_sender_mail,
            self.create_job2_mail,
        ]
        return random.choice(creators)()


    def main(self) -> None:
        self.create_example_attachment_files()

        print("FakeEmailjobsGenerator: i'm alive")

# to create fake scheduled jobs
class FakeSchedulejobsGenerator:

    def add_random_row(self, path="Example_ERP_table.xlsx") -> str:

        if not os.path.isfile(path):
            raise RuntimeError("Example_ERP_table.xlsx not found, run main.py first ")
        wb = load_workbook(path)
        ws = wb.active

        assert ws is not None

        next_row = ws.max_row + 1

        erp_order_number = str(random.randint(10000000, 10999999))
        order_qty = random.randint(10, 100) * 100

        material_available = order_qty + random.randint(-100, 100)

        ws[f"A{next_row}"] = erp_order_number
        ws[f"B{next_row}"] = order_qty
        ws[f"C{next_row}"] = material_available

        wb.save(path)
        wb.close()
        return erp_order_number

# produce random job from above
class FakeJobsGenerator:
    def __init__(self) -> None:

        self.fake_emailjob = FakeEmailjobsGenerator()
        self.fake_scheduledjob = FakeSchedulejobsGenerator()


    def run(self):


        while True:
            try:
                input("\nHit Enter to generate an random job")
                if random.randint(0,1) == 1:
                    path = self.fake_emailjob.create_random_mail()
                    print(f"Created emailjob: {path.name}")
                else:
                    erp_order_number = self.fake_scheduledjob.add_random_row()
                    print(f"Created scheduledjob: {erp_order_number}")

            except KeyboardInterrupt:
                print("\nStopped.")
                break
            except Exception as err:
                print(f"WARN: generator error: {err}")
                time.sleep(1)


# to simualte the the behaviour of the RPA tool implementation 
class RPAToolSimulator:

    def run(self):

        self.log_system("RPAToolSimulator: i'm alive")
        print("RPAToolSimulator: i'm alive")
        
        while True:

            time.sleep(1)

            # read handover
            with open("handover.json", "r", encoding="utf-8") as f:
                handover_data = json.load(f)

            # claim workflow if "job_queued"
            ipc_state = handover_data.get("ipc_state")
            if ipc_state != "job_queued":
                continue

            # singal to Orchestrator the workflow is claimed 
            handover_data["ipc_state"] = "job_running"
            with open("handover.json", "w", encoding="utf-8") as f:
                json.dump(handover_data, f, indent=2)
            
            # identify job type
            job_type = handover_data.get("job_type")
            job_id = handover_data.get("job_id")

            
            #time.sleep(2) # simulate some time... 

            # ----------------------------------------------
            # JOB1
            # ----------------------------------------------
            if job_type == "job1":
                # retrive job-specific data  
                rpa_payload = handover_data.get("rpa_payload", {})
                erp_order_number = rpa_payload.get("order_number")  
                new_qty = rpa_payload.get("target_order_qty")

                # simulation of job1 screenactiviy
                self.log_system(f"activities on screen_1 in ERP completed", job_id)
                self.log_system(f"activities on screen_2 in ERP completed", job_id)
                self.simulate_RPA_result_job1(erp_order_number, new_qty)

                # ready to verify result
                handover_data["ipc_state"] = "job_verifying"
            
            # ----------------------------------------------
            # JOB3
            # ----------------------------------------------
            elif job_type == "job3":
                # retrive job-specific data
                rpa_payload = handover_data.get("rpa_payload", {})
                erp_order_number = rpa_payload.get("source_ref")  
                new_qty = rpa_payload.get("target_order_qty")

                #simulation of job3 screenactiviy
                self.log_system(f"activities on screen_1 in ERP completed", job_id)
                self.log_system(f"activities on screen_2 in ERP completed", job_id)
                self.simulate_RPA_result_job1(erp_order_number, new_qty) # change to a "job3"-activity

                # ready to verify result
                handover_data["ipc_state"] = "job_verifying"

            # ----------------------------------------------
            # PING
            # ----------------------------------------------
            elif job_type == "ping":

                # play a sound
                if platform.system() == "Windows":
                    import winsound
                    winsound.Beep(1000, 300) #type: ignore

                elif platform.system() == "Linux":
                    print("\a", end="", flush=True)

                self.log_system(f"made a ping", job_id)


                # ready to verify result
                handover_data["ipc_state"] = "job_verifying"
            
            # ----------------------------------------------
            # UNKOWN JOB
            # ----------------------------------------------
            else:
                self.log_system(f"no logic for job_type{job_type}")

                # error signal
                handover_data["ipc_state"] = "safestop"


            # ----------------------------------------------
            # Handover to Orchestrator
            # ----------------------------------------------
            with open("handover.json", "w", encoding="utf-8") as f:
                json.dump(handover_data, f, indent=2)

            self.log_system(f"RPASimulator.run() done, ipc_state: job_running -> job_verifying", handover_data.get("job_id"))


    def log_system(self, event_text: str, job_id=None):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
   
        log_line = f"{timestamp} | rpa | job_id={job_id or ''} | RPAToolSimulator...() | {event_text}"

        with open("system.log", "a", encoding="utf-8") as f:
            f.write(log_line + "\n")
            f.flush()
         


    

    def simulate_RPA_result_job1(self, erp_order_number: str, new_qty: int, path="Example_ERP_table.xlsx"):
        # here, updating a row.  IRL: updating in ERP

        assert erp_order_number is not None
        assert new_qty is not None

        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(erp_order_number):
                row[1].value = int(new_qty) #updating to the new value 'in ERP'     # type: ignore
                wb.save(path)
                wb.close()
                return True


def main():

    if not os.path.isfile("main.py"):
        raise RuntimeError("Place this file in main.py directory")

    #rpa_tool_simulator = RPAToolSimulator()
    #threading.Thread(target=rpa_tool_simulator.run, daemon=True).start() #replace with RPA tool
    
    time.sleep(0.01)

    fakejobs_generator = FakeJobsGenerator() # replace with real sources
    fakejobs_generator.run()

main()