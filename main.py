import atexit
import datetime
import json
import os
import platform
import re
import shutil
import signal
import sqlite3
import subprocess
import sys
import tempfile
import threading
import time
import traceback
import tkinter as tk
from dataclasses import asdict, dataclass
from email import policy
from email.parser import BytesParser
from email.utils import parseaddr
from pathlib import Path
from typing import Any, Literal, Never, TypeAlias, get_args
from zipfile import BadZipFile

from openpyxl import Workbook, load_workbook  # type: ignore

VERSION = "0.4"

# =========================
# CUSTOMIZATION POINTS
# =========================
# Replace or adapt these parts for a real environment:
# - ExampleMailBackend
# - ExampleErpBackend
# - ExampleJob*Handler classes
# - NetworkService.NETWORK_HEALTHCHECK_PATH
# - FriendsRepository file format, if needed
# - the external RPA tool that reads/writes handover.json (not included here)


# ============================================================
# DATA MODELS
# ============================================================

IpcState: TypeAlias = Literal["idle", "job_queued", "job_running", "job_verifying", "safestop"]
JobType: TypeAlias = Literal["ping", "job1", "job2", "job3", "job4"]
JobSourceType: TypeAlias = Literal["personal_inbox", "shared_inbox", "erp_query"]
JobStatus: TypeAlias = Literal["REJECTED", "QUEUED", "RUNNING", "VERIFYING", "FAIL", "DONE"]
JobAction: TypeAlias = Literal["DELETE_ONLY", "REPLY_AND_DELETE", "QUEUE_RPA_TOOL", "SKIP", "MOVE_BACK_TO_INBOX"]
UIStatusText: TypeAlias = Literal["online", "safestop", "working", "no network" , "ooo"]
UserOutcome: TypeAlias = Literal["DONE", "REJECTED", "PRE_HANDOVER_CRASH", "RPA_TOOL_CRASH", "VERIFICATION_MISMATCH", "POST_HANDOVER_CRASH", "OUT_OF_SERVICE",]

@dataclass
class JobCandidate:
    source_ref: str
    job_source_type: JobSourceType
    source_data: dict[str, Any]

    sender_email: str | None = None # for email only
    subject: str | None = None  # for email only
    body: str | None = None  # for email only


@dataclass
class JobDecision:
    action: JobAction
    job_type: JobType | None = None
    job_status: JobStatus | None = None
    error_code: str | None = None
    error_message: str | None = None
    rpatool_payload: dict[str, Any] | None = None
    ui_log_message: str | None = None
    system_log_message: str | None = None
    send_lifesign_notice: bool = False
    start_recording: bool = False


@dataclass
class ActiveJob:
    """active_job transfers to handover.json and is excanges with the RPA tool."""
    
    # common fields
    ipc_state: IpcState

    source_ref: str | None = None  # identifier, eg. "ERP_ORDER:12345" or "mail1234.eml"

    job_type: JobType | None = None
    job_source_type: JobSourceType | None = None
    job_id: int | None = None

    sender_email: str | None = None # for email
    subject: str | None = None      # for email
    body: str | None = None         # for email eg. "Hi, change the order 12345 to 44 pcs"

    # parsed from source 
    source_data: dict[str, Any] | None = None # eg. {"order_number": 12345, "target_qty": 44}

    # final instruction to RPA tool
    rpatool_payload: dict[str, Any] | None = None # eg. {"order_number": 12345, "target_qty": 44, "pick_qty_from_location": "WH7",}


@dataclass
class PollResult:
    handled_anything: bool
    active_job: ActiveJob | None = None



# ============================================================
# EXAMPLE BACKENDS
# ============================================================

class ExampleMailBackend:
    """
    Example mailbox backend that simulates mailbox processing using local
    folders and .eml files.

    Replace this with a real backend, for example Outlook or Microsoft Graph.
    """


    def __init__(self, logger, job_source_type) -> None:
        self.logger = logger
        self.job_source_type = job_source_type # change to folder in e.g. outlook
        self.inbox_dir = Path(self.job_source_type) / "inbox"
        self.processing_dir = Path(self.job_source_type) / "processing"

        self.inbox_dir.mkdir(parents=True, exist_ok=True)
        self.processing_dir.mkdir(parents=True, exist_ok=True)


    def fetch_from_inbox(self, max_items=None) -> list[str]:
        paths_raw = sorted(self.inbox_dir.glob("*.eml"))

        if max_items is not None:
            paths_raw = paths_raw[:max_items]

        paths = [str(x) for x in paths_raw] #convert Path-type to str

        #self.logger.system(f"fetched {paths}")

        return paths
    

    def parse_mail_file(self, processing_path) -> JobCandidate:
        with open(processing_path, "rb") as f:
            msg = BytesParser(policy=policy.default).parse(f)

        from_name, from_address = parseaddr(msg.get("From", ""))
        subject = msg.get("Subject", "").strip()

        del from_name # not used

        # message_id = msg.get("Message-ID", "").strip()
        # not needed. source_ref is sufficient (in this example: Path.   In outlook: Outlook EntryID / Graph ID)

        # raw_headers = {k: str(v) for k, v in msg.items()}   
        # not needed (but good for troubleshooting all metadata) 

        if msg.is_multipart():
            body_parts = []
            for part in msg.walk():
                if part.get_content_type() == "text/plain" and not part.get_filename():
                    try:
                        body_parts.append(part.get_content())
                    except Exception:
                        pass
            body = "\n".join(body_parts).strip()
        else:
            try:
                body = msg.get_content().strip()
            except Exception:
                body = ""
        

        # placeholder for implementation
        attachments = {}
        #attachments = {
        #    "attachments": [
        #        {
        #            "filename": "orders.xlsx",
        #            "path": "/some/path/orders.xlsx",
        #        }
        #    ]
        #}
     
        return JobCandidate(
            source_ref=processing_path,
            sender_email=from_address.strip().lower(),
            subject=subject,
            body=body,
            job_source_type=self.job_source_type,
            source_data=attachments,
            )


    def claim_to_processing(self, mail: JobCandidate) -> JobCandidate:

        example_path = Path(mail.source_ref) # example backend use Path

        target_path = self.processing_dir / example_path.name # .name for filename only
        shutil.move(str(example_path), str(target_path))
        
        self.logger.system(f"moved {example_path} to {target_path}")
        mail.source_ref = str(target_path)

        return mail
        

    def reply_and_delete(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int) -> None:
        self.send_reply(candidate, extra_subject, extra_body, job_id)
        self.delete_from_processing(candidate, job_id)


    def send_reply(self, candidate: JobCandidate, extra_subject: str, extra_body: str, job_id: int) -> None:

        reply_to = candidate.sender_email
        subject = f"{extra_subject} re: {candidate.subject}"
        body = (
            f"{extra_body} \n\n"
            f"-------------------------------------------------------------\n"
            f"{candidate.body}"
        ) # In a real mail backend, this should use the provider's native reply mechanism.

        reply_message = f"reply_to={reply_to}, subject={subject}, body={body}"
        self.logger.system(reply_message[:200], job_id)
        
        assert reply_to is not None
        self.print_email_preview(reply_to, subject, body)


    def print_email_preview(self, reply_to: str, subject: str, body: str):

        print(
        "\n" + "="*72 +
        "\n📧 EMAIL REPLY PREVIEW\n" +
        "="*72 +
        f"\nFrom:    robot@runtime.local"
        f"\nTo:      {reply_to}"
        f"\nSubject: {subject}"
        f"\nDate:    {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        "\n" + "-"*72 +
        f"\n{body}\n" +
        "="*72 + "\n"
    )
        

    def delete_from_processing(self, candidate: JobCandidate, job_id: int | None = None) -> None:

        self.logger.system(f"removing: {candidate.source_ref}", job_id)
        os.remove(candidate.source_ref)


    def move_back_to_inbox(self, candidate: JobCandidate, job_id: int) -> None:
        ''' to simplify for end-user, return unhandled emails to origin location'''
        
        # placeholder for implementation

        # rename email subject to "FAIL/" and ignore this keywork in is_shared_inbox_email_in_scope()

        example_path = Path(candidate.source_ref)

        target_path = self.inbox_dir / example_path.name #.name only the filenamne
        shutil.move(str(example_path), str(target_path))
        
        self.logger.system(f"moved {candidate} back to inbox", job_id)


class ExampleErpBackend:
    """Example ERP backend backed by a local Excel file."""
    def select_all_from_erp(self, path="Example_ERP_table.xlsx") -> list[dict]:
        # do a well targeted 'query' 

        self.ensure_example_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active

        assert ws is not None #to satisfy pylance

        all_rows=[]

        for row in ws.iter_rows(min_row=2):  # skip header
            
            source_ref = row[0].value
            order_qty = row[1].value
            material_available = row[2].value

            if order_qty != material_available:

                all_rows.append({
                        "source_ref": source_ref,
                        "order_qty": order_qty,
                        "material_available": material_available,
                    })
                
        wb.close()
        return all_rows
    
    
    def parse_row(self, row) -> JobCandidate:
              
        source_ref = row.get("source_ref")
        order_qty = row.get("order_qty")
        material_available = row.get("material_available")


        try: order_qty = int(order_qty)
        except Exception: raise ValueError(f"invalid order_qty: {order_qty}")
        try: material_available = int(material_available)
        except Exception: raise ValueError(f"invalid material_available: {material_available}")


        source_data ={
            "order_qty": order_qty,
            "material_available": material_available,
        }

        return JobCandidate(
            source_ref=str(source_ref),
            job_source_type="erp_query",
            source_data=source_data
        )

    
    def ensure_example_erp_exists(self, path="Example_ERP_table.xlsx") -> None:
        ''' a table in ERP '''
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None #to satisfy pylance

        # headers
        ws["A1"] = "source_ref"
        ws["B1"] = "order_qty"
        ws["C1"] = "material_available"

        wb.save(path)
        wb.close()

    
    def get_order_qty(self, source_ref, path="Example_ERP_table.xlsx") -> int | None:
        self.ensure_example_erp_exists(path)

        try:
            wb = load_workbook(path)
        except BadZipFile:
            time.sleep(1)
            wb = load_workbook(path)

        ws = wb.active
        assert ws is not None #to satisfy pylance

        for row in ws.iter_rows(min_row=2):
            cell_source_ref = row[0].value

            if str(cell_source_ref) == str(source_ref):
                value = row[1].value  # order_qty

                if isinstance(value, int):
                    wb.close()
                    return int(value)
                
                else: 
                    raise ValueError(f"order_qty: {value} is not INT")
        
        wb.close()
        return None


# ============================================================
# JOB FLOWS
# ============================================================

class MailFlow:
    ''' handle email events'''
    def __init__(self, logger, friends_repo, is_within_operating_hours, network_service, job_handlers, pre_handover_executor, mail_backend_personal, mail_backend_shared) -> None:
        self.logger = logger
        self.friends_repo = friends_repo
        self.is_within_operating_hours = is_within_operating_hours
        self.network_service = network_service
        self.job_handlers = job_handlers
        self.pre_handover_executor = pre_handover_executor
        self.mail_backend_personal = mail_backend_personal
        self.mail_backend_shared = mail_backend_shared


    def poll_once(self) -> PollResult:
        ''' a candidate is any email from personal_inbox OR an 'in scope'-email from shared_inbox '''

        # claim and parse from all mail-sources
        candidate = self.claim_next_mail_candidate() 
        if not candidate:
            return PollResult(handled_anything=False, active_job=None)
        
        # personal inbox = direct human-to-runtime channel 
        if candidate.job_source_type == "personal_inbox":
            self.friends_repo.reload_if_modified()
            self.logger.ui(f"email from {candidate.sender_email}", blank_line_before=True)
            decision = self.decide_personal_inbox_email(candidate)

        # shared inbox = external business mailbox
        elif candidate.job_source_type == "shared_inbox":
            decision = self.decide_shared_inbox_email(candidate)


        active_job = self.pre_handover_executor.apply_decision(candidate, decision)
        return PollResult(handled_anything=True, active_job=active_job)


    def claim_next_mail_candidate(self) -> JobCandidate | None:

        # personal inbox (parse, always claim)
        paths = self.mail_backend_personal.fetch_from_inbox(max_items=1)
       
        for path in paths:
            mail = self.mail_backend_personal.parse_mail_file(path)
            mail = self.mail_backend_personal.claim_to_processing(mail)
            self.logger.system(f"{mail.job_source_type} produced mail {mail.source_ref}")
            return mail
        
        # shared inbox (parse, maybe claim)
        paths = self.mail_backend_shared.fetch_from_inbox()
        
        for path in paths:
            mail = self.mail_backend_shared.parse_mail_file(path)

            if not self.is_shared_inbox_email_in_scope(mail):
                continue
            
            mail = self.mail_backend_shared.claim_to_processing(mail)
            self.logger.system(f"{mail.job_source_type} produced mail {mail.source_ref}")

            return mail

        return None


    def is_shared_inbox_email_in_scope(self, mail: JobCandidate) -> bool:
  
        # Intentionally minimal example.
        self.logger.system(f"checking sender: {mail.sender_email} subject: {mail.subject}")

        # skip emails moved back by move_back_to_inbox()
        if str(mail.subject).upper().startswith("FAIL/"):
            return False
        
        # Placeholder for mailbox-specific scope rules, for example supplier or subject matching.
        
        return True

    
    def classify_personal_inbox_email(self, mail: JobCandidate) -> JobType | None:

        subject = str(mail.subject).strip().lower()

        if subject == "ping":
            return "ping"
        
        elif "job1" in subject.lower():
            return"job1"
        
        elif "job2" in subject.lower():
            return "job2"

        return None


    def classify_shared_inbox_email(self):
        # placeholder for implementation
        pass
 

    def decide_personal_inbox_email(self, mail: JobCandidate) -> JobDecision:
        '''decide what to do with the found candidate'''
        job_type = None

        try:
            if not self.friends_repo.is_allowed_sender(mail.sender_email):
                return JobDecision(
                    action="DELETE_ONLY",
                    ui_log_message="--> rejected (not in friends.xlsx)",
                    system_log_message="--> rejected (not in friends.xlsx)"
                )

            if not self.is_within_operating_hours():
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_status="REJECTED",
                    error_code="OUTSIDE_WORKING_HOURS",
                    error_message="Outside working hours 05-23.",
                    ui_log_message="--> rejected (outside working hours)",
                    system_log_message="--> rejected (outside working hours)",
                )

            job_type = self.classify_personal_inbox_email(mail)

            if job_type is None:
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=None,
                    job_status="REJECTED",
                    error_code="UNKNOWN_JOB",
                    error_message="Could not identify a job type.",
                    ui_log_message="--> rejected (unable to identify job type)",
                    system_log_message="--> rejected (unable to identify job type)",
                )
            

            if not self.friends_repo.has_job_access(mail.sender_email, job_type):
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="NO_ACCESS",
                    error_message=f"No access to {job_type}",
                    ui_log_message=f"--> rejected (no access to {job_type})",
                    system_log_message=f"--> rejected (no access to {job_type})",
                )


            if not self.network_service.has_network_access():
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="NO_NETWORK",
                    error_message="No network connection. Your email was removed.",
                    ui_log_message="--> rejected (no network connection)",
                    system_log_message="--> rejected (no network connection)",
                )

            handler = self.job_handlers.get(job_type)
            if handler is None:
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_status="REJECTED",
                    job_type=job_type,
                    error_code="PRE_HANDOVER_ERROR",
                    error_message=f"No handler registered for job_type={job_type}",
                )

            ok, payload_or_error = handler.precheck_and_build_payload(mail)
            if not ok:
                error = str(payload_or_error)
                return JobDecision(
                    action="REPLY_AND_DELETE",
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="INVALID_INPUT",
                    error_message=error,
                    ui_log_message=f"--> rejected (invalid input for {job_type})",
                    system_log_message=f"--> rejected (invalid input for {job_type})",
                )
            
            rpatool_payload = payload_or_error

            return JobDecision(
                action="QUEUE_RPA_TOOL",
                job_type=job_type,
                job_status="QUEUED",
                system_log_message=f"accepted ({job_type})",
                send_lifesign_notice=True,
                start_recording=True,
                rpatool_payload=rpatool_payload
            )

        except Exception as err:
            return JobDecision(
                action="REPLY_AND_DELETE",
                job_status="REJECTED",
                job_type=None,
                error_code="PRE_HANDOVER_ERROR",
                error_message=f"unhandeled MailFlow error: {err}",
            )
    

    def decide_shared_inbox_email(self, mail: JobCandidate) -> JobDecision:
        ''' decide what to do with the found candidate '''
        # placeholder for implementation

        return JobDecision(
                    action="MOVE_BACK_TO_INBOX",
                    job_type=None,
                    error_code="NO_LOGIC",
                    error_message="No logic implemented yet to handle shared_inbox mails",
                    system_log_message="No logic implemented yet to handle shared_inbox mails",
                )


class QueryFlow:
    ''' handle query events'''
    def __init__(self, logger, audit_repo, job_handlers, pre_handover_executor, is_within_operating_hours, erp_backend) -> None:
        self.logger = logger
        self.audit_repo = audit_repo
        self.job_handlers = job_handlers
        self.pre_handover_executor = pre_handover_executor
        self.is_within_operating_hours = is_within_operating_hours
        self.erp_backend = erp_backend


    def poll_once(self) -> PollResult:
        # find candidates in form of a query row 
        
        if not self.is_within_operating_hours():
            return PollResult(handled_anything=False, active_job=None)

        # check if new match
        candidate = self.fetch_next_query_candidate()
        if not candidate:
            return PollResult(handled_anything=False, active_job=None)

        self.logger.ui(f"query job detected for {candidate.source_ref}", blank_line_before=True)
        
        # decide what to do
        decision = self.decide_candidate(candidate)
        
        # do it
        active_job = self.pre_handover_executor.apply_decision(candidate, decision)
        
        return PollResult(handled_anything=True, active_job=active_job)
    
        
    def fetch_next_query_candidate(self) -> JobCandidate | None:

        # job 3 example
        all_selected_rows = self.erp_backend.select_all_from_erp()
        
        if not all_selected_rows:
            return None
    
        for row_candidate_raw in all_selected_rows:
            row_candidate = self.erp_backend.parse_row(row_candidate_raw)

            # Avoid reprocessing the same source_ref multiple times on the same day.
            if self.audit_repo.has_been_processed_today(row_candidate.source_ref):
                continue

            row_candidate.job_source_type="erp_query"
            self.logger.system(f"{row_candidate.job_source_type} produced source_ref {row_candidate.source_ref}")
            return row_candidate
        
        # job 4
        # placeholder for implementation
        
        return None


    def decide_candidate(self, candidate: JobCandidate) -> JobDecision:
        self.logger.system("running")

        job_type = None

        try:

            # placeholder for implementation
            # eg. below:


            job_type = self.classify_candidate(candidate)
            handler = self.job_handlers.get(job_type)

            if handler is None:
                return JobDecision(
                    action="SKIP",
                    job_status="REJECTED",
                    job_type=job_type,
                    error_code="PRE_HANDOVER_ERROR",
                    error_message=f"No handler found for job_type={job_type}",
                )
            
            # do the precheck from "job specifics"-section
            ok, payload_or_error = handler.precheck_and_build_payload(candidate)

            if not ok:
                error = str(payload_or_error)
                return JobDecision(
                    action="SKIP",  
                    job_type=job_type,
                    job_status="REJECTED",
                    error_code="INVALID_INPUT",
                    error_message=error,
                    ui_log_message=f"--> rejected (invalid input for {job_type})",
                )

            rpatool_payload = payload_or_error
            
            
            return JobDecision(
                action="QUEUE_RPA_TOOL",
                job_type=job_type,
                job_status="QUEUED",
                system_log_message=f"accepted ({job_type})",
                start_recording=True,
                rpatool_payload=rpatool_payload,
                )

        except Exception as err:
            return JobDecision(
                action="SKIP",
                job_status="REJECTED",
                job_type=job_type,
                error_code="PRE_HANDOVER_ERROR",
                error_message=str(err),
            )
    

    def classify_candidate(self, candidate: JobCandidate) -> JobType | None:
        # placeholder for implementation

        return "job3"


class PreHandoverExecutor:
    """Execute pre-handover actions and build ActiveJob objects for the RPA tool."""
    def __init__(self, logger, update_ui_status, show_recording_overlay, generate_job_id, recording_service, audit_repo, notification_service, mail_backend_personal, mail_backend_shared) -> None:
        self.logger = logger
        self.recording_service = recording_service
        self.generate_job_id = generate_job_id
        self.audit_repo = audit_repo
        self.update_ui_status = update_ui_status
        self.show_recording_overlay = show_recording_overlay
        self.notification_service = notification_service
        self.mail_backend_personal = mail_backend_personal
        self.mail_backend_shared = mail_backend_shared
    

    def validate_decision(self, decision: JobDecision) -> None:

        if not isinstance(decision, JobDecision):
            raise ValueError("decision must be JobDecision")

        # validate action
        if decision.action not in get_args(JobAction):
            raise ValueError(f"invalid action: {decision.action}")

        # validate job_type
        if decision.job_type is not None and decision.job_type not in get_args(JobType):
            raise ValueError(f"invalid job_type: {decision.job_type}")

        # validate job_status
        if decision.job_status is not None and decision.job_status not in get_args(JobStatus):
            raise ValueError(f"invalid job_status: {decision.job_status}")

        # ============================================================
        # ACTION-SPECIFIC RULES
        # ============================================================

        action = decision.action

        # ------------------------------------------------------------
        # DELETE_ONLY
        # ------------------------------------------------------------
        if action == "DELETE_ONLY":
            if decision.job_status is not None:
                raise ValueError("DELETE_ONLY must not have job_status")
            if decision.rpatool_payload is not None:
                raise ValueError("DELETE_ONLY must not have rpatool_payload")

        # ------------------------------------------------------------
        # REPLY_AND_DELETE
        # ------------------------------------------------------------
        elif action == "REPLY_AND_DELETE":
            if decision.job_status != "REJECTED":
                raise ValueError("REPLY_AND_DELETE requires job_status REJECTED")

            if decision.error_message is None:
                raise ValueError("REPLY_AND_DELETE requires error_message (reply text)")

            if decision.rpatool_payload is not None:
                raise ValueError("REPLY_AND_DELETE must not have rpatool_payload")
   
        # ------------------------------------------------------------
        # QUEUE_RPA_TOOL
        # ------------------------------------------------------------
        elif action == "QUEUE_RPA_TOOL":
            if decision.job_type is None:
                raise ValueError("QUEUE_RPA_TOOL requires job_type")

            if decision.job_status != "QUEUED":
                raise ValueError("QUEUE_RPA_TOOL requires job_status='QUEUED'")

            if decision.rpatool_payload is None:
                raise ValueError("QUEUE_RPA_TOOL requires rpatool_payload")
            

            if not isinstance(decision.rpatool_payload, dict):
                raise ValueError("rpatool_payload must be dict")

        # ------------------------------------------------------------
        # SKIP (query flow reject)
        # ------------------------------------------------------------
        elif action == "SKIP":
            if decision.job_status != "REJECTED":
                raise ValueError("SKIP requires job_status='REJECTED'")

        # ------------------------------------------------------------
        # MOVE_BACK_TO_INBOX
        # ------------------------------------------------------------
        elif action == "MOVE_BACK_TO_INBOX":
            if decision.job_status is not None:
                raise ValueError("MOVE_BACK_TO_INBOX should not set job_status")

        else:
            # should never happen due to earlier validation
            raise ValueError(f"Unhandled action: {action}")
        

    def validate_candidate_decision_combination(self, candidate: JobCandidate, decision: JobDecision) -> None:

        if candidate.job_source_type not in get_args(JobSourceType):
            raise ValueError(f"invalid candidate.job_source_type: {candidate.job_source_type}")

        is_mail = candidate.job_source_type in ("personal_inbox", "shared_inbox")
        is_query = candidate.job_source_type == "erp_query"

        # ------------------------------------------------------------
        # source-type-specific candidate sanity
        # ------------------------------------------------------------
        if is_mail:
            if candidate.sender_email is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires sender_email")
            if candidate.subject is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires subject")
            if candidate.body is None:
                raise ValueError(f"{candidate.job_source_type} candidate requires body")

        elif is_query:
            if candidate.source_data is None:
                raise ValueError("erp_query candidate requires source_data")

        else:
            raise ValueError(f"unknown candidate.job_source_type: {candidate.job_source_type}")

        # ------------------------------------------------------------
        # action must match candidate source type
        # ------------------------------------------------------------
        if decision.action in ("DELETE_ONLY", "REPLY_AND_DELETE", "MOVE_BACK_TO_INBOX"):
            if not is_mail:
                raise ValueError(
                    f"action {decision.action} is only valid for mail candidates, not {candidate.job_source_type}"
                )

        if decision.action == "SKIP":
            if not is_query:
                raise ValueError(
                    f"action SKIP is only valid for query candidates, not {candidate.job_source_type}"
                )

        # ------------------------------------------------------------
        # optional policy checks
        # ------------------------------------------------------------
        if decision.send_lifesign_notice:
            if candidate.job_source_type != "personal_inbox":
                raise ValueError("lifesign only valid for personal_inbox")

            if not candidate.sender_email:
                raise ValueError("lifesign requires sender_email")
            
        # MOVE_BACK_TO_INBOX is for shared-only emails
        if decision.action == "MOVE_BACK_TO_INBOX":
            if candidate.job_source_type != "shared_inbox":
                raise ValueError("MOVE_BACK_TO_INBOX is only valid for shared_inbox")

        # DELETE_ONLY is for personal-only emails.
        if decision.action == "DELETE_ONLY":
            if candidate.job_source_type != "personal_inbox":
                raise ValueError("DELETE_ONLY is only valid for personal_inbox")


        # REPLY_AND_DELETE should be used for personal inbox only 
        if decision.action == "REPLY_AND_DELETE" and candidate.job_source_type != "personal_inbox":
            raise ValueError("REPLY_AND_DELETE is only valid for personal_inbox")

        # send_lifesign_notice only makes sense together with QUEUE_RPA_TOOL
        if decision.send_lifesign_notice and decision.action != "QUEUE_RPA_TOOL":
            raise ValueError("send_lifesign_notice requires action='QUEUE_RPA_TOOL'")

        # start_recording only makes sense for queued jobs
        if decision.start_recording and decision.action != "QUEUE_RPA_TOOL":
            raise ValueError("start_recording requires action='QUEUE_RPA_TOOL'")


    def _log_decision_messages(self, decision: JobDecision):
        if decision.ui_log_message:
            self.logger.ui(decision.ui_log_message)

        if decision.system_log_message:
            self.logger.system(decision.system_log_message)


    def _maybe_send_lifesign(self, candidate: JobCandidate, decision: JobDecision, job_id: int|None):
        if decision.send_lifesign_notice and not self.audit_repo.has_sender_job_today(candidate.sender_email, job_id):
            self.notification_service.send_lifesign(candidate, job_id)


    def _maybe_start_recording(self, decision: JobDecision, job_id: int|None):
        
        if decision.start_recording:
            self.recording_service.start(job_id)
            
            try: self.show_recording_overlay()
            except Exception as e: self.logger.system(f"error {e}", job_id)


    def _insert_audit_row(self, candidate: JobCandidate, decision: JobDecision, job_id: int|None,):
        now = datetime.datetime.now()
        job_finish_time = None if decision.action == "QUEUE_RPA_TOOL" else now.strftime("%H:%M:%S")

        self.audit_repo.insert_job(
        job_id=job_id,
        source_ref=candidate.source_ref,
        email_address=candidate.sender_email,
        email_subject=candidate.subject,
        job_type=decision.job_type,
        job_start_date=now.strftime("%Y-%m-%d"),
        job_start_time=now.strftime("%H:%M:%S"),
        job_finish_time=job_finish_time,
        job_status=decision.job_status,
        job_source_type=candidate.job_source_type,
        error_code=decision.error_code,
        error_message=decision.error_message,
        )


    def _build_active_job(self, candidate: JobCandidate, decision: JobDecision, job_id: int| None) -> ActiveJob:
        
        return ActiveJob(
            ipc_state="job_queued",
            job_id=job_id,
            job_type=decision.job_type,
            job_source_type=candidate.job_source_type,
            source_ref=candidate.source_ref,
            sender_email=candidate.sender_email,
            subject=candidate.subject,
            body=candidate.body,
            source_data=candidate.source_data,
            rpatool_payload=decision.rpatool_payload,
            )
    

    def apply_decision(self, candidate: JobCandidate, decision: JobDecision) -> ActiveJob | None:
        
        self.validate_decision(decision)
        self.validate_candidate_decision_combination(candidate, decision)

        self._log_decision_messages(decision)
        
        # DELETE_ONLY is intentionally non-audited.
        if decision.action == "DELETE_ONLY": # only personal-inbox e.g. not in friends.xlsx
            self.mail_backend_personal.delete_from_processing(candidate)
            return
        
        job_id = self.generate_job_id()
        try:
            # prioritize new audit row
            self._insert_audit_row(candidate, decision, job_id,)

            # REPLY_AND_DELETE
            if decision.action == "REPLY_AND_DELETE": # only personal-inbox
                self.notification_service.send_job_reply(
                    candidate=candidate,
                    outcome="REJECTED",
                    job_id=job_id,
                    reason=decision.error_message,
                )
                self.audit_repo.update_job(job_id=job_id, final_reply_sent=True,)
                return
            
            # SKIP
            if decision.action == "SKIP":
                return

            # MOVE_BACK_TO_INBOX
            elif decision.action == "MOVE_BACK_TO_INBOX": # only shared-inbox e.g. error with email in scope
                self.mail_backend_shared.move_back_to_inbox(candidate, job_id)
                return
                            
            # QUEUE_RPA_TOOL
            if decision.action == "QUEUE_RPA_TOOL":
                self.update_ui_status(forced_status="working")
                self._maybe_send_lifesign(candidate, decision, job_id)
                self._maybe_start_recording(decision, job_id)
                active_job = self._build_active_job(candidate, decision, job_id)
                return active_job


        # allow job_id to bubble up to degraded_mode for error in _insert_audit_row(), _maybe_send_lifesign(), _maybe_start_recording
        except Exception as e:
            raise RuntimeError(f"PRE_HANDOVER_CRASH: {e} crash_job_id={job_id}")


class PostHandoverFinalizer:
    """Finalize jobs after the RPA tool returns control. Verification is cold-start based."""
    def __init__(self, logger, audit_repo, job_handlers, recording_service, hide_recording_overlay, mail_backend_personal, mail_backend_shared, notification_service) -> None:
        self.logger = logger
        self.audit_repo = audit_repo
        self.job_handlers = job_handlers
        self.recording_service = recording_service
        self.hide_recording_overlay = hide_recording_overlay
        self.mail_backend_personal = mail_backend_personal
        self.mail_backend_shared = mail_backend_shared
        self.notification_service = notification_service


    def poll_once(self, active_job: ActiveJob, job_id: int) -> None:
        try:
            # note in audit that the job is 're-taken' from RPA
            self.audit_repo.update_job(job_id=job_id, job_status="VERIFYING")
            
            job_type = active_job.job_type
            handler = self.job_handlers.get(job_type)
            if handler is None:
                ok_or_error = f"No handler for job_type={job_type}"
            else:
                ok_or_error = handler.verify_result(active_job)

            if ok_or_error == "ok":
                self.finalize_job_result(active_job, job_status="DONE", error_code=None, error_message=None)
                return
            
        except Exception as e:
             # verification stage crashed, outcome unknown
            try:
                self.audit_repo.update_job(
                job_id=job_id,
                job_status="FAIL",
                error_code="POST_HANDOVER_CRASH",
                error_message=f"crash during verification stage: {e}",
                job_finish_time=datetime.datetime.now().strftime("%H:%M:%S"),
            )
            except Exception as e2:
                self.logger.system(f"[PostHandoverFinalizer] {e} {e2}", job_id)
                

            raise RuntimeError(f"{e} (POST_HANDOVER_CRASH)")
        

        if str(ok_or_error).endswith("(VERIFICATION_MISMATCH)"):
        # consider mismatch a critical error 
            self.finalize_job_result(
                active_job,
                job_status="FAIL",
                error_code="VERIFICATION_MISMATCH",
                error_message=str(ok_or_error),)
            raise RuntimeError(f"{ok_or_error}")
        
        raise RuntimeError(f"Unknown ok_or_error: {ok_or_error} (POST_HANDOVER_CRASH)")


    def _update_audit(self, job_id, job_status, error_code, jobhandler_error_message,) -> None:
        
        now = datetime.datetime.now().strftime("%H:%M:%S")
        
        self.audit_repo.update_job(
            job_id=job_id, 
            job_status=job_status, 
            error_code=error_code, 
            error_message=jobhandler_error_message, 
            job_finish_time=now,
            )


    def _update_logs(self, job_status: str, active_job: ActiveJob,) -> None:
        
        job_status = job_status.lower()
        job_type = active_job.job_type

        # ui/dashboard log
        self.logger.ui(f"--> {job_status} ({job_type})")

        # system log (system.log)
        self.logger.system(f"{job_status} ({job_type})", active_job.job_id)

        
    def finalize_job_result(self, active_job: ActiveJob, job_status: JobStatus, error_code: str | None, error_message: str | None):
        job_id = active_job.job_id

        # update audit w/ result (DONE/FAIL)
        self._update_audit(job_id, job_status, error_code, error_message)
        
        # do side effects
        self.recording_service.stop(job_id)
        self.recording_service.upload_recording(job_id)
        self.hide_recording_overlay()
 
        # do mail source specifics (email)
        final_reply_sent = self.handle_source_completion(active_job, job_status, error_code, error_message)

        if final_reply_sent:
            self.audit_repo.update_job(job_id=job_id, final_reply_sent=True,)

        # update ui w/ result (DONE/FAIL)
        self._update_logs(job_status, active_job)


     
    def _map_candidate_from_activejob(self, active_job: ActiveJob) -> JobCandidate:
        assert active_job.source_ref is not None # to satisfy pylance
        assert active_job.source_data is not None # to satisfy pylance
        assert active_job.job_source_type is not None # to satisfy pylance
        
        return JobCandidate(
                source_ref=active_job.source_ref,
                job_source_type=active_job.job_source_type,
                source_data=active_job.source_data,
                sender_email=active_job.sender_email,
                subject=active_job.subject,
                body=active_job.body,
                )
        
  
    def handle_source_completion(self, active_job: ActiveJob, job_status: str, error_code: str | None, error_message: str | None) -> bool:
        if active_job.job_source_type == "erp_query":
            return False

        if active_job.job_source_type not in ("personal_inbox", "shared_inbox"):
            return False
    
        # rebuild candidate (from rebuilt active_job)
        candidate = self._map_candidate_from_activejob(active_job)

        job_id = active_job.job_id

        # send final reply (and delete)
        if active_job.job_source_type == "personal_inbox":
            if job_status == "DONE":
                self.notification_service.send_job_reply(
                    candidate=candidate,
                    outcome="DONE",
                    job_id=job_id,
                )
                return True

            elif job_status == "FAIL":
                if error_code == "VERIFICATION_MISMATCH":
                    outcome = "VERIFICATION_MISMATCH"
                elif error_code == "POST_HANDOVER_CRASH":
                    outcome = "POST_HANDOVER_CRASH"
                else:
                    outcome = "RPA_TOOL_CRASH"  # eller raise, om du vill vara strikt

                self.notification_service.send_job_reply(
                    candidate=candidate,
                    outcome=outcome,
                    job_id=job_id,
                    reason=error_message,
                )
                return True

            raise ValueError(f"unexpected job_status in handle_source_completion(): {job_status}")

        # delete shared mail (or move to archive?)
        if active_job.job_source_type == "shared_inbox":
            self.mail_backend_shared.delete_from_processing(candidate, job_id)
            return False
        
        return False


# ============================================================
# JOB SPECIFICS
# ============================================================

class ExampleJob1Handler:
    ''' everything for job1 '''
    def __init__(self, logger) -> None:
        self.logger = logger


    def precheck_and_build_payload(self, candidate: JobCandidate) -> tuple[bool, dict | str]:
        ''' sanity-check (and ERP check) on given data '''
        body = candidate.body
        assert body is not None # to satisfy pylance

        # get important info for job1, eg.:
        order_number_match = re.search(r"order_number:\s*(.+)", body)
        order_number = order_number_match.group(1) if order_number_match else None

        order_qty_match = re.search(r"order_qty:\s*(.+)", body)
        order_qty = order_qty_match.group(1) if order_qty_match else None

        material_available_match = re.search(r"material_available:\s*(.+)", body)
        material_available = material_available_match.group(1) if material_available_match else None

        error = ""
        if order_number is None:
            error += "missing order_number. "
        if order_qty is None:
            error += "missing order_qty. "
        if material_available is None:
            error += "missing material_available. "

        if error:
            return False, error.strip()

        # and for any attachments, eg:
        attachments = candidate.source_data.get("attachments", [])
        #for attachment in attachments:
        #    print(attachment.get("filename"))

        rpatool_payload = {
            "order_number": order_number,
            "order_qty": order_qty,
            "target_order_qty": material_available,
            "attachments": attachments,
        }

        return True, rpatool_payload
    

    def verify_result(self, activejob: ActiveJob):
        return "ok"


class ExampleJob2Handler:
    ''' everything for job2 '''
    def __init__(self, logger) -> None:
        self.logger = logger
   

    def precheck_and_build_payload(self, candidate: JobCandidate) -> tuple[bool, dict | str]:
        # placeholder for implementation

        return False, "no logic for job2."


    def verify_result(self, activejob: ActiveJob):
        return "ok"


class ExamplePingJobHandler:
    ''' everything for ping '''
    def __init__(self, logger) -> None:
        self.logger = logger


    def precheck_and_build_payload(self, candidate: JobCandidate) -> tuple[bool, dict | str]:
        return True, {}


    def verify_result(self, activejob: ActiveJob):
        return "ok"
    
   
class ExampleJob3Handler:
    ''' everything for job3 '''
    def __init__(self, logger, erp_backend) -> None:
        self.logger = logger
        self.erp_backend = erp_backend

   
    def precheck_and_build_payload(self, candidate: JobCandidate) -> tuple[bool, dict[str, Any] | str]:
        source_ref = candidate.source_ref
        order_qty = candidate.source_data.get("order_qty")
        material_available = candidate.source_data.get("material_available")

        if order_qty == material_available:
            return False, "no mismatch left to fix"

        rpatool_payload = {
            "source_ref": str(source_ref),
            "target_order_qty": material_available,
        }

        return True, rpatool_payload
    

    def verify_result(self, active_job: ActiveJob) -> str:
    
        job_id = active_job.job_id

        # get erp order number/id
        rpatool_payload = active_job.rpatool_payload
        if not rpatool_payload:
            return "missing rpatool_payload"
        
        # get the order number/id and the target qty sent to RPA tool
        source_ref = rpatool_payload.get("source_ref")
        target_order_qty = rpatool_payload.get("target_order_qty")


        # get the 'real' qty now in erp
        order_qty_erp = self.erp_backend.get_order_qty(source_ref)

        # compare them
        if order_qty_erp != target_order_qty:
            message= f"ERP shows mismatch. {source_ref} should be {target_order_qty}, is {order_qty_erp} (VERIFICATION_MISMATCH)"
            self.logger.system(message, job_id)
            return message

        self.logger.system(f"OK. Should be: {target_order_qty}, is: {order_qty_erp}", job_id)
        return "ok"



# ============================================================
# HANDOVER / IPC
# ============================================================

class HandoverRepository:
    """Persist and validate the file-based IPC state shared with the RPA tool."""
    def __init__(self, logger) -> None:
        self.logger = logger
        self.HANDOVER_FILE = "handover.json"

    def read(self) -> ActiveJob:
        ''' read HANDOVER_FILE '''
        
        last_err=None

        for attempt in range(7):
            try:
                # read file
                with open(self.HANDOVER_FILE, "r", encoding="utf-8") as f:
                    handover_data = json.load(f)
                
                # rebuild object
                active_job = self.validate_and_build_activejob(handover_data)

                return active_job
                
            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: retry {attempt+1}/7 : {err}")
                time.sleep(attempt/10)
        
        
        raise RuntimeError(f"{self.HANDOVER_FILE} unreadable: {last_err}")
    
      
    def write(self, active_job: ActiveJob) -> None:
        ''' atomic write of HANDOVER_FILE '''

        handover_data = asdict(active_job)

        self.validate_and_build_activejob(handover_data) # only validate (ignore return)
        job_id = handover_data.get("job_id")

        last_err = None
        
        for attempt in range(7):
            temp_path = None
            try:
                
                dir_path = os.path.dirname(os.path.abspath(self.HANDOVER_FILE))
                fd, temp_path = tempfile.mkstemp(dir=dir_path, suffix=".tmp")

                #atomic write
                with os.fdopen(fd, "w", encoding="utf-8") as tmp:
                    json.dump(handover_data, tmp, indent=2) # indent for human eyes
                    tmp.flush()
                    os.fsync(tmp.fileno())

                os.replace(temp_path, self.HANDOVER_FILE)
                
                self.logger.system(
                    f"prepared handover for job_type {handover_data.get('job_type')} "
                    f"with rpatool_payload {handover_data.get('rpatool_payload')}",
                    job_id,
                )               
                return

            except Exception as err:
                last_err = err
                self.logger.system(f"WARN: {attempt+1}/7 error", job_id)
                time.sleep(attempt/10) # 0 0.1... 0.6 sec     

            finally:
                if temp_path and os.path.exists(temp_path):
                    try: os.remove(temp_path)
                    except Exception: pass

        self.logger.system(f"CRITICAL: cannot write {self.HANDOVER_FILE} {last_err}", job_id)
        raise RuntimeError(f"CRITICAL: cannot write {self.HANDOVER_FILE}")


    def validate_and_build_activejob(self, handover_data: dict) -> ActiveJob:
        """Validate raw handover dict and return ActiveJob."""

        ipc_state = handover_data.get("ipc_state")
        job_id = handover_data.get("job_id")
        job_type = handover_data.get("job_type")
        job_source_type = handover_data.get("job_source_type")
        source_ref = handover_data.get("source_ref")
        sender_email = handover_data.get("sender_email")
        subject = handover_data.get("subject")
        body = handover_data.get("body")
        source_data = handover_data.get("source_data")
        rpatool_payload = handover_data.get("rpatool_payload")

        if ipc_state is None:
            raise ValueError("ipc_state missing")

        if ipc_state not in get_args(IpcState):
            raise ValueError(f"unknown state: {ipc_state}")

        if job_id is not None:
            try:
                job_id = int(job_id)
            except Exception:
                raise ValueError(f"job_id not INT-like: {job_id}")

        if ipc_state == "idle":
            if any(v is not None for v in (
                job_id, job_type, job_source_type, source_ref,
                sender_email, subject, body, source_data, rpatool_payload
            )):
                raise ValueError(f"state 'idle' should have no more variables: {handover_data}")

        elif ipc_state in ("job_queued", "job_running", "job_verifying"):
            required_fields = {
                "job_id": job_id,
                "job_type": job_type,
                "job_source_type": job_source_type,
                "source_ref": source_ref,
                "source_data": source_data,
                "rpatool_payload": rpatool_payload,
            }

            missing = [k for k, v in required_fields.items() if v is None]
            if missing:
                raise ValueError(f"{ipc_state} has missing fields in {self.HANDOVER_FILE}: {missing}")

            if job_type not in get_args(JobType):
                raise ValueError(f"unknown job_type: {job_type}")

            if job_source_type not in get_args(JobSourceType):
                raise ValueError(f"unknown job_source_type: {job_source_type}")

            if job_source_type in ("personal_inbox", "shared_inbox"):
                required_fields = {
                    "sender_email": sender_email,
                    "subject": subject,
                    "body": body,
                }

                missing = [k for k, v in required_fields.items() if v is None]
                if missing:
                    raise ValueError(f"{job_source_type} has missing fields in {self.HANDOVER_FILE}: {missing}")
                
            if not isinstance(source_data, dict):
                raise ValueError("source_data must be dict")
            if not isinstance(rpatool_payload, dict):
                raise ValueError("rpatool_payload must be dict")

        return ActiveJob(
            ipc_state=ipc_state,
            job_id=job_id,
            job_type=job_type,
            job_source_type=job_source_type,
            source_ref=source_ref,
            sender_email=sender_email,
            subject=subject,
            body=body,
            source_data=source_data,
            rpatool_payload=rpatool_payload,
        )


    def is_valid_ipc_transition(self, prev_ipc_state: IpcState | None, ipc_state: IpcState) -> bool:
        """ transition-validator for RobotRuntime loop. Only runs when ipc_state != prev_ipc_state. """

        if prev_ipc_state is None: # at startup
            return True

        allowed_transitions: dict[IpcState, set[IpcState]] = {
            "idle": {"job_queued", "safestop"},
            "job_queued": {"job_running", "safestop"},
            "job_running": {"job_verifying", "safestop"},
            "job_verifying": {"idle", "safestop"},
            "safestop": {"idle"},
        }

        allowed_next = allowed_transitions[prev_ipc_state]

        if ipc_state not in allowed_next:
            return False
        
        return True


# ============================================================
# USER NOTIFICATIONS
# ============================================================

class UserNotificationService:
    """Only for personal_inbox user-facing replies."""

    ADMIN_EMAIL = "admin_rpa-AT-company.com"
    COMMAND_JOB_ID = 999999999999

    def __init__(self, mail_backend_personal, recordings_destination_folder: str, watchdog_timeout: int):
        self.mail_backend_personal = mail_backend_personal
        self.recordings_destination_folder = recordings_destination_folder
        self.watchdog_timeout = watchdog_timeout

    def send_job_reply(self, candidate: JobCandidate, outcome: UserOutcome, job_id: int, reason=None, from_safestop:bool=False, from_initialize:bool=False) -> None:
        
        subject, body = self._build_job_reply(
            outcome=outcome,
            job_id=job_id,
            reason=reason,
            from_safestop=from_safestop,
            from_initialize=from_initialize,
        )

        self._send(candidate, subject, body, job_id, delete_after=True)

    def send_recovery_reply(self, audit_row: dict, candidate: JobCandidate, from_safestop:bool, from_initialize:bool) -> None:
        job_id = audit_row["job_id"]
        job_status = audit_row["job_status"]
        error_code = audit_row.get("error_code")
        error_message = audit_row.get("error_message")


        outcome: UserOutcome
        if job_status == "DONE":
            outcome = "DONE"
        elif job_status == "REJECTED":
            outcome = "REJECTED"
        elif job_status == "QUEUED":
            outcome = "PRE_HANDOVER_CRASH"
        elif job_status == "RUNNING":
            outcome = "RPA_TOOL_CRASH"
        elif job_status == "VERIFYING":
            outcome = "POST_HANDOVER_CRASH"
        elif job_status == "FAIL":
            if error_code == "PRE_HANDOVER_CRASH":
                outcome = "PRE_HANDOVER_CRASH"
            elif error_code == "RPA_TOOL_CRASH":
                outcome = "RPA_TOOL_CRASH"
            elif error_code == "VERIFICATION_MISMATCH":
                outcome = "VERIFICATION_MISMATCH"
            elif error_code == "POST_HANDOVER_CRASH":
                outcome = "POST_HANDOVER_CRASH"
            else:
                outcome = "RPA_TOOL_CRASH"
        else:
            outcome = "RPA_TOOL_CRASH"

        self.send_job_reply(
            candidate=candidate,
            outcome=outcome,
            job_id=job_id,
            reason=error_message,
            from_safestop=from_safestop,
            from_initialize=from_initialize,
        )

    def send_out_of_service_reply(self, candidate: JobCandidate, job_id: int) -> None:
        self.send_job_reply(
            candidate=candidate,
            outcome="OUT_OF_SERVICE",
            job_id=job_id,
            reason="Robot is out-of-service and does not accept any new jobs.",
        )

    def send_command_reply(self, candidate: JobCandidate) -> None:
        self._send(
            candidate=candidate,
            subject="got it!",
            body="Command received.",
            job_id=self.COMMAND_JOB_ID,
            delete_after=True,
        )

    def send_admin_alert(self, reason: str) -> None:
        fake_candidate = JobCandidate(
            source_ref="safestop, no real source_ref",
            sender_email=self.ADMIN_EMAIL,
            subject="",
            body="",
            job_source_type="personal_inbox",
            source_data={},
        )

        body = (
            "Robot is in degraded mode.\n\n"
            f"Reason:\n{reason}\n\n"
            "Available commands: 'stop1234' and 'restart1234'."
        )

        self._send(
            candidate=fake_candidate,
            subject="safestop notice",
            body=body,
            job_id=self.COMMAND_JOB_ID,
            delete_after=False,
        )

    def send_lifesign(self, candidate: JobCandidate, job_id: int) -> None:
        body = (
            ">Hello, human<\n\n"
            "The first request each day is replied with: online\n"
            "Next message is sent after completion\n"
            f"(in max {self.watchdog_timeout} seconds from now).\n"
        )

        self._send(
            candidate=candidate,
            subject="ONLINE",
            body=body,
            job_id=job_id,
            delete_after=False,
        )


    def _build_job_reply(self, outcome: UserOutcome, job_id: int, reason, from_safestop:bool, from_initialize:bool) -> tuple[str, str]:
        # note to self: for increased user value, extend reply with a short summary of result, eg. "changed PO 450221 on SKU 110212 from 34pcs to 31pcs"
        recording_text = self._get_recording_text(job_id)

        subject: str
        body: str

        if outcome == "DONE":
            subject = "DONE"
            body = (
                "Job completed successfully.\n\n"
                f"Job ID: {job_id}\n\n"
                f"{recording_text}"
                "This email can be deleted."
            )


        elif outcome == "PRE_HANDOVER_CRASH" or outcome == "REJECTED":
            subject = "FAIL"
            body = (
                    "Your request was not started.\n\n"
                    f"{self._format_reason(reason)}"
                    f"Job ID: {job_id}\n"
                    "Keep calm, no changes were made in ERP.\n\n"
                    )
            if from_safestop:
                body += (
                    "To avoid further problems, the robot will go out-of-service.\n"
                )
            body += (
                    "No action is required from your side.\n"
                    "This email can be deleted."
                )

        elif outcome == "RPA_TOOL_CRASH":
            subject = "FAIL"
            body = (
                    "The robot started your request, but then crashed.\n\n"
                    f"{self._format_reason(reason)}"
                    f"Job ID: {job_id}\n"
                    "Changes may have been made in ERP before the crash.\n"
                    "It is (very) recommended that you review the result manually.\n\n"
                    f"{recording_text}"
                )
            if from_safestop:
                body += (
                    "To avoid further problems, the robot will go out-of-service.\n"
                )  
            body += (
                    "This email can be deleted."
                )
                

        elif outcome == "VERIFICATION_MISMATCH":
            subject = "FAIL"
            body = (
                    "The robot completed the request, and the result was checked in ERP.\n"
                    "However, the final ERP data did not match the expected result.\n\n"
                    f"{self._format_reason(reason)}"
                    f"Job ID: {job_id}\n"
                    "You NEED TO review the result manually in ERP.\n\n"
                    f"{recording_text}"
                    )
            if from_safestop:
                body += (
                    "To avoid further problems, the robot will go out-of-service.\n"
                )  
            body += (
                    "This email can be deleted."
                )
        
        elif outcome == "POST_HANDOVER_CRASH":
            subject = "FAIL"
            body = (
                    "The robot completed the request, but crashed during the final verification stage.\n"
                    "The outcome could therefore not be confirmed automatically.\n\n"
                    f"{self._format_reason(reason)}"
                    f"Job ID: {job_id}\n"
                    "Please verify the result manually in ERP.\n\n"
                    f"{recording_text}"
                    )
            if from_safestop:
                body += (
                    "To avoid further problems, the robot will go out-of-service.\n"
                )  
            body += (
                    "This email can be deleted."
                )

        elif outcome == "OUT_OF_SERVICE":
            subject = "FAIL"
            body = (
                "The robot is out-of-service and does not accept new jobs.\n\n"
                #f"{self._format_reason(reason)}"
                #f"Job ID: {job_id}\n"
                #"No changes were made in ERP.\n\n"
                "This email can be deleted."
            )

        else:
            raise ValueError(f"Unknown outcome: {outcome}")

        if from_initialize:
            body = (
                "The robot crashed and has now restarted.\n"
                "If you already received a final reply (DONE/FAIL) for this job, you can ignore this recovery message."
            ) + "\n\n" + body

        return subject, body

    def _get_recording_text(self, job_id: int) -> str:
        recording_path = self._get_recording_path(job_id)
        if not recording_path:
            return ""

        return (
            "A screen recording is available for review:\n"
            f"{recording_path}\n\n"
        )

    def _get_recording_path(self, job_id: int) -> str | None:
        path = Path(self.recordings_destination_folder) / f"{job_id}.mkv"
        if path.exists():
            return str(path)
        return None

    def _format_reason(self, reason: str | None) -> str:
        if not reason:
            return ""
        return f"Reason: {reason}\n\n"

    def _send(self, candidate: JobCandidate, subject: str, body: str, job_id: int, delete_after: bool,) -> None:
        if delete_after:
            self.mail_backend_personal.reply_and_delete(
                candidate=candidate,
                extra_subject=subject,
                extra_body=body,
                job_id=job_id,
            )
        else:
            self.mail_backend_personal.send_reply(
                candidate=candidate,
                extra_subject=subject,
                extra_body=body,
                job_id=job_id,
            )


# ============================================================
# RECORDING / SAFESTOP / INFRASTRUCTURE
# ============================================================   
                      
class RecordingService:
    ''' screen-recording to capture all RPA tool screen-activity '''

    RECORDINGS_IN_PROGRESS_FOLDER = "recordings_in_progress"
    RECORDINGS_DESTINATION_FOLDER = "recordings_destination"

    def __init__(self, logger,) -> None:
        #written by AI

        self.logger = logger
        self.recording_process = None


    def get_screen_resolution(self):
        try:
            output = subprocess.check_output(["xrandr"], text=True)
            for line in output.splitlines():
                if "*" in line:
                    res = line.split()[0]
                    return res.split("x")
        except Exception:
            pass

        # fallback: Tkinter
        try:
            root = tk.Tk()
            root.withdraw()
            width = root.winfo_screenwidth()
            height = root.winfo_screenheight()
            root.destroy()
            return str(width), str(height)
        except Exception:
            pass

        return "1920", "1080"

 
    def start(self, job_id) -> None:
        """start the screen recording"""
        # written by AI

        os.makedirs(self.RECORDINGS_IN_PROGRESS_FOLDER, exist_ok=True)
        filename = f"{self.RECORDINGS_IN_PROGRESS_FOLDER}/{job_id}.mkv"

        drawtext = (
            f"drawtext=text='job_id  {job_id}':"
            "x=200:y=20:"
            "fontsize=32:"
            "fontcolor=lightyellow:"
            "box=1:"
            "boxcolor=black@0.5"
        )

        if platform.system() == "Windows":
            ffmpeg_path = None

            local_ffmpeg = Path("./ffmpeg.exe")
            if local_ffmpeg.exists():
                ffmpeg_path = str(local_ffmpeg)
            else:
                ffmpeg_in_path = shutil.which("ffmpeg")
                if ffmpeg_in_path:
                    ffmpeg_path = ffmpeg_in_path

            if ffmpeg_path is None:
                message = (
                    "WARN: screen-recording disabled because ffmpeg was not found. "
                    "Place ffmpeg.exe next to main.py or install ffmpeg in PATH."
                )
                print(message)
                self.logger.system(message, job_id)
                return

            capture = ["-f", "gdigrab", "-i", "desktop"]

            recording_process = subprocess.Popen(
                [
                    ffmpeg_path,
                    "-y",
                    *capture,
                    "-framerate", "15",
                    "-vf", drawtext,
                    "-vcodec", "libx264",
                    "-preset", "ultrafast",
                    filename,
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0),
            )

        else:
            display = os.environ.get("DISPLAY")
            if not display:
                self.logger.system("WARN: screen-recording disabled because DISPLAY is missing", job_id)
                return

            ffmpeg_path = shutil.which("ffmpeg")
            if ffmpeg_path is None:
                self.logger.system("WARN: screen-recording disabled because ffmpeg is not installed", job_id)
                return

            width, height = self.get_screen_resolution()

            capture = [
                "-video_size", f"{width}x{height}",
                "-f", "x11grab",
                "-i", display,
            ]

            recording_process = subprocess.Popen(
                [
                    ffmpeg_path,
                    "-y",
                    *capture,
                    "-framerate", "15",
                    "-vf", drawtext,
                    "-vcodec", "libx264",
                    "-preset", "ultrafast",
                    filename,
                ],
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                start_new_session=True,
            )

        time.sleep(0.2)

        if recording_process.poll() is not None:
            self.logger.system("WARN: ffmpeg exited immediately; recording did not start", job_id)
            return

        self.recording_process = recording_process
        self.logger.system("recording started", job_id)

        
    def stop(self, job_id=None) -> None:
        ''' allow global kill of FFMPEG processes since Orchestrator is designed to run on a dedicated machine '''
        # written by AI

        try: self.logger.system("stop recording", job_id)
        except Exception: pass

        recording_process = self.recording_process
        self.recording_process = None

        try:
            if recording_process is not None:
                # try first stop only our own process
                if platform.system() == "Windows":
                    try:
                        recording_process.send_signal(
                            getattr(signal, "CTRL_BREAK_EVENT", signal.SIGTERM)
                        )
                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        subprocess.run(
                            ["taskkill", "/PID", str(recording_process.pid), "/T", "/F"],
                            stdout=subprocess.DEVNULL,
                            stderr=subprocess.DEVNULL,
                            check=False,
                        )
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

                else:
                    # try first stop only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGINT)

                    except Exception:
                        try:
                            recording_process.terminate()
                        except Exception:
                            pass

                    try:
                        recording_process.wait(timeout=8)
                        return
                    except subprocess.TimeoutExpired:
                        pass

                    # else, kill only our own process
                    try:
                        os.killpg(recording_process.pid, signal.SIGKILL)
                        recording_process.wait(timeout=3)
                        return
                    except Exception:
                        pass

                    # last resort, global kill all ffmpeg
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )

            else:
                # fallback if process object is lost
                if platform.system() == "Windows":
                    subprocess.run(
                        ["taskkill", "/IM", "ffmpeg.exe", "/T", "/F"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
                else:
                    subprocess.run(
                        ["killall", "-q", "-KILL", "ffmpeg"],
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL,
                        check=False,
                    )
        except Exception as err:
            self.logger.system(f"WARN from stop(): {err}", job_id)


    def upload_recording(self, job_id, max_attempts=3) -> None:
        ''' upload to a shared drive'''
    
        local_file = f"{self.RECORDINGS_IN_PROGRESS_FOLDER}/{job_id}.mkv"
        local_file = Path(local_file)

        remote_path = Path(self.RECORDINGS_DESTINATION_FOLDER) / f"{job_id}.mkv"
        remote_path.parent.mkdir(parents=True, exist_ok=True)

        for attempt in range(max_attempts):
            try:
                
                shutil.copy2(local_file, remote_path)
                self.logger.system(f"upload success: {remote_path}", job_id)
                try: os.remove(local_file)
                except Exception: pass

                return

            except Exception as e:
                self.logger.system(f"Attempt {attempt+1}/{max_attempts} failed: {e}", job_id)
                time.sleep(attempt + 1)
        
        self.logger.system(f"upload failed: {remote_path}", job_id)


    def cleanup_aborted_recordings(self):
        ''' cleanup aborted screen-recordings '''

        directory = Path(self.RECORDINGS_IN_PROGRESS_FOLDER)
        if not directory.exists():
            return
        
        for file in directory.iterdir():

            if file.is_file() and file.suffix == ".mkv":
                job_id = file.stem
                
                try:
                    self.logger.system(f"cleanup upload started")
                    self.upload_recording(job_id)
                except Exception as err:
                    self.logger.system(f"cleanup failed for {job_id}: {err}")


class FriendsRepository:
    '''Example access-control source for personal_inbox'''

    def __init__(self) -> None:
        self.access_by_email: dict[str, set[str]] = {}
        self.access_file_mtime: float | None = None
        self._lock = threading.Lock()


    def ensure_friends_file_exists(self, path: str = "friends.xlsx") -> None:
        '''Create a template friends.xlsx if missing.'''
        if os.path.exists(path):
            return

        wb = Workbook()
        ws = wb.active
        assert ws is not None

        ws["A1"] = "email"
        ws["B1"] = "ping"
        ws["C1"] = "job1"
        ws["D1"] = "job2"

        ws["A2"] = "alice@example.com"
        ws["B2"] = "x"

        ws["A3"] = "bob@test.com"
        ws["B3"] = "x"
        ws["C3"] = "x"
        ws["D3"] = "x"

        wb.save(path)
        wb.close()


    def _load_access_file(self, filepath: str) -> dict[str, set[str]]:
        '''
        Reads friends.xlsx and returns for example:

        {
            "alice@example.com": {"ping"},
            "ex2@whatever.com": {"ping", "job1"}
        }
        '''
        # written by AI

        wb = load_workbook(filepath, data_only=True)
        try:
            ws = wb.active
            assert ws is not None

            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                raise ValueError("friends.xlsx contains no users")

            header = rows[0]
            self.validate_friends_header(header)
            access_map: dict[str, set[str]] = {}

            for row in rows[1:]:
                email_cell = row[0]
                if email_cell is None:
                    continue

                email = str(email_cell).strip().lower()
                if not email:
                    continue

                permissions: set[str] = set()

                for col in range(1, len(header)):
                    jobname = header[col]
                    if jobname is None:
                        continue

                    jobname = str(jobname).strip().lower()
                    cell = row[col] if col < len(row) else None

                    if cell is None:
                        continue

                    if str(cell).strip().lower() == "x":
                        permissions.add(jobname)

                access_map[email] = permissions

            return access_map
        finally:
            wb.close()


    def reload_if_modified(self) -> bool:
        '''Reload friends.xlsx if changed.'''
        # written by AI

        path = "friends.xlsx"
        self.ensure_friends_file_exists(path)

        mtime = os.path.getmtime(path)
        if self.access_file_mtime == mtime:
            return False

        new_access = self._load_access_file(path)
        self.validate_friends_access(new_access)

        with self._lock:
            self.access_by_email = new_access
            self.access_file_mtime = mtime

        return True


    def is_allowed_sender(self, email_address: str) -> bool:
        email = email_address.strip().lower()
        return email in self.access_by_email


    def has_job_access(self, email_address: str, job_type: str) -> bool:
        email = email_address.strip().lower()
        job = job_type.strip().lower()
        return job in self.access_by_email.get(email, set())


    def validate_friends_access(self, access_map: dict[str, set[str]]) -> None:
        ''' not implemented '''
        if not isinstance(access_map, dict):
            raise ValueError("access_map must be dict")

        valid_job_types = set(get_args(JobType))

        for email, permissions in access_map.items():
            if not isinstance(email, str):
                raise ValueError(f"invalid email key type: {email}")

            email_normalized = email.strip().lower()
            if not email_normalized:
                raise ValueError("empty email in access_map")

            if "@" not in email_normalized:
                raise ValueError(f"invalid email in friends.xlsx: {email}")

            if not isinstance(permissions, set):
                raise ValueError(f"permissions must be set for {email}")

            invalid_permissions = permissions - valid_job_types
            if invalid_permissions:
                raise ValueError(
                    f"invalid job types for {email}: {sorted(invalid_permissions)}. "
                    f"Allowed: {sorted(valid_job_types)}"
                )
            

    def validate_friends_header(self, header_row) -> None:
        ''' validate headers '''
        if not header_row or str(header_row[0]).strip().lower() != "email":
            raise ValueError("friends.xlsx column A must be 'email'")

        valid_job_types = set(get_args(JobType))

        for col in range(1, len(header_row)):
            jobname = header_row[col]
            if jobname is None:
                continue

            jobname_str = str(jobname).strip().lower()
            if jobname_str not in valid_job_types:
                raise ValueError(
                    f"invalid job type column in friends.xlsx: {jobname_str}. "
                    f"Allowed: {sorted(valid_job_types)}"
                )


class NetworkService:
    ''' checks if the computer is connected to company LAN '''
    # Placeholder for implementation
    
    # e.g. NETWORK_HEALTHCHECK_PATH=    r"G:\\"    or    r"\\\\server\\share"
    NETWORK_HEALTHCHECK_PATH = r"/"

    def __init__(self, logger) -> None:
        self.logger = logger
        self.network_state = False #assume offline at start
        self.next_network_check_time = 0


    def has_network_access(self) -> bool:
        #this runs at highest once every hour (if online), or before new jobs

        now = time.time()

        if now < self.next_network_check_time:
            return self.network_state

        try:
            os.listdir(self.NETWORK_HEALTHCHECK_PATH)
            online = True
            
        except Exception:
            online = False
            
        # update log if any network change (and UI? )
        if online != self.network_state:
            self.network_state = online

            if online:
                self.logger.system("network restored")
            else:
                self.logger.system(f"WARN: network lost")

        # check every minute if offline, else every hour
        if online:
            self.next_network_check_time = now + 3600   # 1 h
        else:
            self.next_network_check_time = now + 60     # 1 min
        
        return online


class AuditRepository:
    ''' handles job_audit.db, an audit-style activity log '''

    DB_PATH = "job_audit.db"

    def __init__(self, logger) -> None:
        self.logger = logger

    def _connect(self,) -> sqlite3.Connection:
        conn = sqlite3.connect(self.DB_PATH, timeout=10,)
        return conn   

    def ensure_db_exists(self) -> None:
        
        with self._connect() as conn:
            cur = conn.cursor()
           
            cur.execute('''
                CREATE TABLE IF NOT EXISTS audit_log
                         (
                        job_id INTEGER PRIMARY KEY, 
                        job_type TEXT, 
                        job_status TEXT, 
                        email_address TEXT, 
                        email_subject TEXT, 
                        source_ref TEXT,
                        job_start_date TEXT, 
                        job_start_time TEXT, 
                        job_finish_time TEXT, 
                        final_reply_sent INTEGER NOT NULL DEFAULT 0,
                        job_source_type TEXT,
                        error_code TEXT, 
                        error_message TEXT 
                        )
                        ''')


    def _build_audit_fields(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None = None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None = None, final_reply_sent=None, job_source_type: JobSourceType | None = None, error_code=None, error_message=None,) -> dict:
        all_fields = {
            "job_id": job_id,
            "email_address": email_address,
            "email_subject": email_subject,
            "source_ref": source_ref,
            "job_type": job_type,
            "job_start_date": job_start_date,
            "job_start_time": job_start_time,
            "job_finish_time": job_finish_time,
            "job_status": job_status,
            "final_reply_sent": final_reply_sent,
            "job_source_type": job_source_type,
            "error_code": error_code,
            "error_message": error_message,
        }

        # drop None:s
        fields = {k: v for k, v in all_fields.items() if v is not None}

        self.logger.system(f"received {fields}", job_id)

    
        return fields


    def insert_job(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None=None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None=None, final_reply_sent=None, job_source_type:JobSourceType | None=None, error_code=None, error_message=None,) -> None:
        # use for new row

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_type=job_type,
            job_start_date=job_start_date,
            job_start_time=job_start_time,
            job_finish_time=job_finish_time,
            job_status=job_status,
            final_reply_sent=final_reply_sent,
            job_source_type=job_source_type,
            error_code=error_code,
            error_message=error_message,
        )
        
        columns = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)


        with self._connect() as conn:
            cur = conn.cursor()

            cur.execute(
                f"INSERT INTO audit_log ({columns}) VALUES ({placeholders})",
                tuple(fields.values())
            )


    def update_job(self, job_id, email_address=None, email_subject=None, source_ref=None, job_type: JobType | None=None, job_start_date=None, job_start_time=None, job_finish_time=None, job_status: JobStatus | None=None, final_reply_sent=None, job_source_type:JobSourceType | None=None, error_code=None, error_message=None,) -> None:
        # example use: self.audit_repo.update_job(job_id=20260311124501, job_type="job1")

        fields = self._build_audit_fields(
            job_id=job_id,
            email_address=email_address,
            email_subject=email_subject,
            source_ref=source_ref,
            job_type=job_type,
            job_start_date=job_start_date,
            job_start_time=job_start_time,
            job_finish_time=job_finish_time,
            job_status=job_status,
            final_reply_sent=final_reply_sent,
            job_source_type=job_source_type,
            error_code=error_code,
            error_message=error_message,
        )
        
        fields.pop("job_id", None)

        if not fields:
            return

        set_clause = ", ".join(f"{k}=?" for k in fields)

        with self._connect() as conn:
            cur = conn.cursor()

            cur.execute(
                f"UPDATE audit_log SET {set_clause} WHERE job_id=?",
                (*fields.values(), job_id)
            )

            if cur.rowcount == 0:
                raise ValueError(f"update_job(): no row in DB with job_id={job_id}")


    def count_done_jobs_today(self) -> int:

        today = datetime.date.today().isoformat()

        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ?
                AND job_status = 'DONE'
            ''', (today,))
            
            result = cur.fetchone()[0]

        return result


    def has_sender_job_today(self, sender_mail, job_id) -> bool:

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ? AND email_address = ? AND job_id != ?
                ''',
                (today, sender_mail, job_id,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0


    def has_been_processed_today(self, source_ref) -> bool:
        # use to avoid bad loops in query-jobs

        today = datetime.datetime.now().strftime("%Y-%m-%d")
        with self._connect() as conn:
            cur = conn.cursor()

            cur.execute(
                '''
                SELECT COUNT(*)
                FROM audit_log
                WHERE job_start_date = ? AND source_ref = ?
                ''',
                (today, source_ref,)
            )

            jobs_today = cur.fetchone()[0]

        return jobs_today > 0


    def get_latest_job_id(self) -> int:
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id
                FROM audit_log
                ORDER BY job_id DESC
                LIMIT 1
            ''')
            row = cur.fetchone()

        return row[0] if row is not None else 0


    def get_failed_jobs(self, days=7):
        ''' not implemented '''
        with self._connect() as conn:
            cur = conn.cursor()
            cur.execute('''
                SELECT job_id, email_address, job_type, error_code, error_message
                FROM audit_log
                WHERE job_status = 'FAIL'
                AND job_start_date >= date('now', '-' || ? || ' days')
                ORDER BY job_id DESC
            ''', (days,))
        res = cur.fetchall()
        
        return res


    def get_pending_reply_jobs(self) -> list[dict]:
        job_source_type: JobSourceType = "personal_inbox"

        with self._connect() as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute(
                '''
                SELECT job_id, email_address, email_subject, source_ref, job_status, error_code, error_message
                FROM audit_log
                WHERE job_source_type = ?
                AND COALESCE(final_reply_sent, 0) = 0
                ORDER BY job_id
                ''',
                (job_source_type,)
            )
            rows = cur.fetchall()

        list_of_dicts = [dict(row) for row in rows]

        return list_of_dicts


class LoggerService:
    """ logging functions"""
    def __init__(self, dashboard_ui) -> None:
        self.dashboard_ui = dashboard_ui

    def ui(self, text:str, blank_line_before: bool = False) -> None:
        
        self.dashboard_ui.post_log_line(text, blank_line_before)


    def system(self, event_text, job_id: int | None=None,):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        event_text = str(event_text)

        # get caller function name
        try:
            frame = sys._getframe(1)
            caller_name = frame.f_code.co_name
            instance = frame.f_locals.get("self")
            if instance is not None:
                class_name = instance.__class__.__name__
                caller = f"{class_name}.{caller_name}()"
            else:
                caller = f"{caller_name}()"

        except Exception:
            caller = "unknown_caller()"
      
        log_line = f"{timestamp} | py  | job_id={job_id or ''} | {caller} | {event_text}"

        # normalize to single-line
        log_line = " ".join(str(log_line).split())

        last_err = None
        for i in range(7):
            try:
                with open("system.log", "a", encoding="utf-8") as f:
                    f.write(log_line + "\n")
                    f.flush()
                return

            except Exception as err:
                last_err = err
                print(f"WARN: retry {i+1}/7 from log_system():", err)
                time.sleep(i + 1)

        # fallback to print() when log fails        
        print(f"[print fallback] {job_id} {event_text} | {last_err}")  
 

class SafeStopController:
    """Handle degraded mode, crash recovery, and operator restart/stop commands."""
    def __init__(self, logger, recording_service, hide_recording_overlay, post_status_update, set_ui_shutdown, mail_backend_personal, audit_repo, generate_job_id, friends_repo, notification_service, poll_for_rpatool_stop_flag) -> None:
        self.logger = logger
        self.recording_service = recording_service
        self.hide_recording_overlay = hide_recording_overlay
        self.post_status_update = post_status_update
        self.set_ui_shutdown = set_ui_shutdown
        self.mail_backend_personal = mail_backend_personal
        self.audit_repo = audit_repo
        self.generate_job_id = generate_job_id
        self.friends_repo = friends_repo
        self._degraded_mode_entered = False
        self.notification_service = notification_service
        self.poll_for_rpatool_stop_flag = poll_for_rpatool_stop_flag


    def run_degraded_mode(self, err_short: str, err_with_traceback: str, active_job:ActiveJob|None) -> None:
        '''
        Rules:
        * no job intake
        * mail-flow inactivated
        * query-flow inactivated
        * no handover to RPA tool
        * 'safestop' status text in UI
        * STOP and RESTART commands allowed 
        * send rejected reply to email from users in friends.xlsx
        * warning email sent to admin
        '''
        
        if self._degraded_mode_entered: return
        self._degraded_mode_entered = True

        if active_job is None or active_job.ipc_state == "idle":
            err_with_traceback_ext = f"ROBOTRUNTIME CRASHED:\n\n{err_with_traceback}"
        
        elif active_job.ipc_state=="job_queued":
        # stop RPA_tool from claiming the workflow
            try:
                active_job.ipc_state="safestop"
                handover_data = asdict(active_job)
                with open("handover.json", "w", encoding="utf-8") as f:
                    json.dump(handover_data, f, indent=2)
            except Exception:
                try: os.remove("handover.json")
                except Exception as e: self.logger.system(e)
            
            err_with_traceback_ext = (
                f"ROBOTRUNTIME CRASHED:\n\n job_type={active_job.job_type} with rpatool_payload {active_job.rpatool_payload}\n\n {err_with_traceback}"
            )

        else:
            err_with_traceback_ext = (
                f"ROBOTRUNTIME CRASHED:\n\n job_type={active_job.job_type} with rpatool_payload {active_job.rpatool_payload}\n\n {err_with_traceback}"
            )

        print(err_with_traceback_ext)

        # try to recover job_id (if any)
        job_id = active_job.job_id if active_job is not None else None
        if job_id is None:
            if "crash_job_id=" in err_short:
                try: job_id = err_short.split("crash_job_id=")[1].split()[0]
                except Exception: pass
     
        self.logger.system(err_with_traceback_ext, job_id)
        
        if job_id is not None:
            error_code = self._classify_crash_error(err_short, active_job)
            try: self.audit_repo.update_job(job_id=job_id, job_status="FAIL", error_code=error_code, error_message=err_short)
            except Exception as e: self.logger.system(e, job_id)

        try: self.recording_service.stop()
        except Exception as e: self.logger.system(e, job_id)

        try: self.recording_service.cleanup_aborted_recordings()
        except Exception as e: self.logger.system(e, job_id)

        try: self.notification_service.send_admin_alert(err_with_traceback)
        except Exception as e: self.logger.system(e, job_id)

        try:
            self.logger.ui("CRASH! All automations halted. Admin is notified.", blank_line_before=True)
            self.logger.ui(f"Reason: {err_short}")
        except Exception as e: self.logger.system(e, job_id)

        try:
            if self.audit_repo.get_pending_reply_jobs():
                try: self.send_recovery_replies(from_safestop=True)
                except Exception as e: self.logger.system(e, job_id)
        except Exception as e: self.logger.system(e, job_id)

        # placeholder to implement recovery reply for emails stuck in 'processing' folder in personal_inbox 

        # placeholder for recovery logic for post_handover crash/mismatch for query jobs

        try: self.hide_recording_overlay()
        except Exception as e: self.logger.system(e, job_id)

        try: self.post_status_update("safestop")
        except Exception as e:
            self.logger.system(e, job_id)
            try: self.set_ui_shutdown()
            except Exception as e2: 
                self.logger.system(e2, job_id)
                os._exit(1)
            time.sleep(3)
            os._exit(0)
        
        self.enter_degraded_loop()


    def _classify_crash_error(self, err_short: str, active_job: ActiveJob|None) -> str:
        if "PRE_HANDOVER_CRASH" in err_short:
            return "PRE_HANDOVER_CRASH"
        if "RPA_TOOL_CRASH" in err_short:
            return "RPA_TOOL_CRASH"
        if "VERIFICATION_MISMATCH" in err_short:
            return "VERIFICATION_MISMATCH"
        if "POST_HANDOVER_CRASH" in err_short:
            return "POST_HANDOVER_CRASH"
        
        if active_job == None:
            return "SAFESTOP"

        if active_job.ipc_state == "job_queued":
            return "PRE_HANDOVER_CRASH"
        if active_job.ipc_state == "job_running":
            return "RPA_TOOL_CRASH"
        if active_job.ipc_state == "job_verifying":
            return "POST_HANDOVER_CRASH"

        return "SAFESTOP"

    def send_recovery_replies(self, from_safestop:bool=False, from_initialize:bool=False) -> None:
        candidate: JobCandidate

        all_jobs = self.audit_repo.get_pending_reply_jobs()

        for audit_row in all_jobs:

            job_id = audit_row.get("job_id")
            source_ref = audit_row.get("source_ref")
            error_code = audit_row.get("error_code")
            
            path = Path(source_ref)
            if not path.exists() and error_code != "RECOVERY_SOURCE_MISSING": # to avoid loop
                self.logger.system(f"recovery skipped: missing processing file {source_ref}", job_id)
                self.audit_repo.update_job(job_id=job_id, error_code="RECOVERY_SOURCE_MISSING")
                continue

            candidate = self.mail_backend_personal.parse_mail_file(str(path))
            self.notification_service.send_recovery_reply(audit_row, candidate, from_safestop, from_initialize)      

            self.logger.system(f"recovery reply sent", job_id)
            
            self.audit_repo.update_job(
                job_id=job_id,
                final_reply_sent=True,
            )


    def _check_for_restart_flag(self, restartflag) -> None:
        if os.path.isfile(restartflag):
            try: os.remove(restartflag)
            except Exception: pass
            self.logger.system(f"restart-command received from {restartflag}")
            self.restart_application()


    def _check_for_restart_command(self, candidate: JobCandidate) -> None:
        if candidate.subject is None:
            return

        if "restart1234" in candidate.subject.strip().lower():
            self.logger.system(f"restart command received from {candidate.sender_email}")
            try: self.notification_service.send_command_reply(candidate)
            except Exception: pass
            self.restart_application()


    def _check_for_stop_command(self, candidate: JobCandidate) -> None:
        if candidate.subject is None:
            return

        if "stop1234" in candidate.subject.strip().lower():
            self.logger.system(f"stop command received from {candidate.sender_email}")
            try: self.notification_service.send_command_reply(candidate)
            except Exception: pass
            try: self.set_ui_shutdown()
            except Exception: os._exit(1)
            os._exit(0)


    def _try_notify_user(self, candidate: JobCandidate, job_id: int) -> bool:
        final_reply_sent = False

        try:
            self.notification_service.send_out_of_service_reply(candidate, job_id)
            final_reply_sent = True

        except Exception as e:
            self.logger.system(e, job_id)
            
        return final_reply_sent


    def _try_insert_rejected_audit(self, job_id:int, candidate:JobCandidate, final_reply_sent: bool):
        try:
            now = datetime.datetime.now()
            job_source_type: JobSourceType = "personal_inbox" 
            
            self.audit_repo.insert_job(
                job_id=job_id,
                source_ref=candidate.source_ref,
                email_address=candidate.sender_email,
                email_subject=candidate.subject,
                job_start_date=now.strftime("%Y-%m-%d"),
                job_start_time=now.strftime("%H:%M:%S"),
                job_status="REJECTED",
                error_code="IN_SAFESTOP",
                error_message="not accepting new jobs in safestop",
                job_source_type = job_source_type,
                final_reply_sent = final_reply_sent,
            )
        except Exception as e:
            self.logger.system(e, job_id)
            

    def enter_degraded_loop(self) -> Never:
        '''Run essentials, where the priority is replying to user emails.'''  

        self.logger.system("running")
        self.friends_repo.reload_if_modified()
        
        while True:
            try:
                time.sleep(1)
                self.poll_for_rpatool_stop_flag()
                self._check_for_restart_flag("restart.flag")

                # process one personal inbox email in degraded mode
                paths = self.mail_backend_personal.fetch_from_inbox(max_items=1)
                if not paths:
                    continue
                
                inbox_path = paths[0]
                candidate = self.mail_backend_personal.parse_mail_file(inbox_path)                
                candidate = self.mail_backend_personal.claim_to_processing(candidate)

                try: self.logger.ui(f"email from {candidate.sender_email}", blank_line_before=True)
                except Exception: pass

                # silent delete non friends
                if not self.friends_repo.is_allowed_sender(candidate.sender_email):
                    self.logger.ui("--> rejected (not in friends.xlsx)")
                    self.mail_backend_personal.delete_from_processing(candidate)
                    continue
                
                # check for email commands
                self._check_for_restart_command(candidate)
                self._check_for_stop_command(candidate)

                # reply, audit-log and delete for friends
                job_id = self.generate_job_id()
                final_reply_sent = self._try_notify_user(candidate, job_id)
                self._try_insert_rejected_audit(job_id, candidate, final_reply_sent)
                
                try: self.logger.ui("--> rejected (safestop)")
                except Exception: pass
            
            except Exception as e:
                self.logger.system(e)


    def restart_application(self) -> Never:
        # written by AI
    
        self.logger.system("restarting application in new visible terminal")
 
        try:
            self.set_ui_shutdown()
        except Exception:
            pass

        try:
            if platform.system() == "Windows":
                # Open a new visible PowerShell window and run the same script
                subprocess.Popen([
                    "powershell",
                    "-NoExit",
                    "-Command",
                    f'& "{sys.executable}" "{os.path.abspath(sys.argv[0])}"'
                ])

            else:
                # Linux: open a new terminal window and run the same script
                python_cmd = f'"{sys.executable}" "{os.path.abspath(sys.argv[0])}"'

                terminal_candidates = [
                    ["gnome-terminal", "--", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xfce4-terminal", "--hold", "-e", python_cmd],
                    ["konsole", "-e", "bash", "-lc", f"{python_cmd}; exec bash"],
                    ["xterm", "-hold", "-e", python_cmd],
                ]

                launched = False
                for cmd in terminal_candidates:
                    try:
                        subprocess.Popen(cmd)
                        launched = True
                        break
                    except FileNotFoundError:
                        continue

                if not launched:
                    raise RuntimeError("No supported terminal emulator found for restart")

        except Exception as e:
            self.logger.system(e)
            os._exit(1)

        time.sleep(3)
        os._exit(0)


# ============================================================
# UI
# ============================================================

class DashboardUI:
    """Tkinter dashboard for runtime status, logs, and operator visibility."""

    # colors
    BG = "#000000"
    TEXT = "#F5F5F5"
    MUTED = "#A0A0A0"
    GREEN = "#22C55E"
    GREEN_2 = "#16A34A"
    GREEN_3 = "#15803D"
    RED = "#DC2626"
    YELLOW = "#FACC15"
    SCROLL_TROUGH = "#0F172A"
    SCROLL_BG = "#1E293B"
    SCROLL_ACTIVE = "#475569"

    # fonts
    FONT_STATUS = ("Arial", 100, "bold")
    FONT_COUNTER = ("Segoe UI", 140, "bold")
    FONT_SMALL = ("Arial", 14, "bold")
    FONT_LOG = ("DejaVu Sans Mono", 20)
    FONT_RECORDING = ("Arial", 20, "bold")

    # sizes
    WINDOW_GEOMETRY = "1800x1000+0+0"
    ROOT_PADX = 50
    SCROLLBAR_WIDTH = 23

    RECORDING_WIDTH = 250
    RECORDING_HEIGHT = 110
    RECORDING_MARGIN_RIGHT = 30


    def __init__(self, shutdown_callback=None):
        self.shutdown_callback = shutdown_callback
        self._build_root(self.BG)
        self._build_header(self.BG, self.TEXT)
        self._build_body(self.BG, self.TEXT)
        self._build_footer(self.BG, self.TEXT)

        #self.debug_grid(self.root)


    def run(self) -> None:
        self.root.mainloop()


    def set_shutdown_callback(self, callback) -> None:
        self.shutdown_callback = callback

    def shutdown(self) -> None:
        if self._closing:
            return

        self._closing = True

        try:
            if self.shutdown_callback is not None:
                self.shutdown_callback()
        except Exception:
            pass

        self.root.destroy()


    def debug_grid(self, widget):
        ''' highlights all grids with red '''
        for child in widget.winfo_children():
            try:
                child.configure(highlightbackground="red", highlightthickness=1)
            except Exception:
                pass
            self.debug_grid(child)


    def _build_root(self, bg_color):
        self.root = tk.Tk()
        self.root.geometry(self.WINDOW_GEOMETRY)
        self.root.resizable(False, False)

        self.root.configure(bg=bg_color, padx=self.ROOT_PADX)
        self._closing = False
        self.root.protocol("WM_DELETE_WINDOW", self.shutdown)

        self.root.title('RPA dashboard')
        self._create_recording_overlay()

        # layout using grid
        self.root.grid_rowconfigure(1, weight=1)
        self.root.grid_columnconfigure(0, weight=1)


    def _build_header(self, bg_color, text_color):
        self.header = tk.Frame(self.root, bg=bg_color)

        self.header.grid(row=0, column=0, sticky="ew")
        self.header.grid_columnconfigure(2, weight=1)
        self.header.grid_rowconfigure(0, weight=1)

        # Header content
        self.rpa_text_label = tk.Label(
            self.header,
            text="RPA:",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_STATUS,
        )  
        self.rpa_text_label.grid(row=0, column=0, padx=16, pady=16, sticky="w")

        self.rpa_status_label = tk.Label(
            self.header,
            text="",
            fg=self.RED,
            bg=bg_color,
            font=self.FONT_STATUS,
        )
        self.rpa_status_label.grid(row=0, column=1, padx=16, pady=16, sticky="w")

        self.status_dot = tk.Label(
            self.header,
            text="",
            fg=self.GREEN,
            bg=bg_color,
            font=("Arial", 50, "bold"),
        )
        self.status_dot.grid(row=0, column=2, sticky="w")

        # jobs done today (counter + label in same grid)
        self.jobs_counter_frame = tk.Frame(self.header, bg=bg_color)
        self.jobs_counter_frame.grid(row=0, column=3, sticky="ne", padx=40, pady=30)
        self.jobs_counter_frame.grid_rowconfigure(0, weight=1)
        self.jobs_counter_frame.grid_columnconfigure(0, weight=1)

        # normal view (jobs done today)
        self.jobs_normal_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_normal_view.grid(row=0, column=0, sticky="nsew")
        self.jobs_normal_view.grid_columnconfigure(0, weight=1)

        self.jobs_done_label = tk.Label(
            self.jobs_normal_view,
            text="0",
            fg=text_color,
            bg=bg_color,
            font=self.FONT_COUNTER,
            anchor="e",
            justify="right",
        )
        self.jobs_done_label.grid(row=0, column=0, sticky="e")

        self.jobs_counter_text = tk.Label(
            self.jobs_normal_view,
            text="jobs done today",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.jobs_counter_text.grid(row=1, column=0, sticky="e", pady=(0, 6))

        # safestop view (big X)
        self.jobs_error_view = tk.Frame(self.jobs_counter_frame, bg=bg_color)
        self.jobs_error_view.grid(row=0, column=0, sticky="nsew")

        self.safestop_x_label = tk.Label(
            self.jobs_error_view,
            text="X",
            bg=self.RED,
            fg="#FFFFFF",
            font=self.FONT_COUNTER,
        )  # text="✖",
        self.safestop_x_label.pack(expand=True)

        # show normal view at startup
        self.jobs_normal_view.tkraise()

        # 'online'-status animation
        self._online_animation_after_id = None
        self._online_pulse_index = 0

        # 'working...'-status animation
        self._working_animation_after_id = None
        self._working_dots = 0


    def _build_body(self, bg_color, text_color):
        self.body = tk.Frame(self.root, bg=bg_color)
        self.body.grid(row=1, column=0, sticky="nsew")
        self.body.grid_rowconfigure(0, weight=1)
        self.body.grid_columnconfigure(0, weight=1)

        # body content
        log_and_scroll_container = tk.Frame(self.body, bg=bg_color)
        log_and_scroll_container.grid(row=0, column=0, sticky="nsew")
        log_and_scroll_container.grid_rowconfigure(0, weight=1)
        log_and_scroll_container.grid_columnconfigure(0, weight=1)

        # the right-hand side scrollbar
        scrollbar = tk.Scrollbar(
            log_and_scroll_container,
            width=self.SCROLLBAR_WIDTH,
            troughcolor=self.SCROLL_TROUGH,
            bg=self.SCROLL_BG,
            activebackground=self.SCROLL_ACTIVE,
            bd=0,
            highlightthickness=0,
            relief="flat",
        )
        scrollbar.grid(row=0, column=1, sticky="ns")

        # the 'console'-style log
        self.log_text = tk.Text(
            log_and_scroll_container,
            yscrollcommand=scrollbar.set,
            bg=bg_color,
            fg=text_color,
            insertbackground="black",
            font=self.FONT_LOG,
            wrap="none",
            state="disabled",
            bd=0,
            highlightthickness=0,
        )  # glow highlightbackground="#1F2937", highlightthickness=1
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.config(command=self.log_text.yview)


    def _build_footer(self, bg_color, text_color):
        self.footer = tk.Frame(self.root, bg=bg_color)
        self.footer.grid(row=2, column=0, sticky="nsew")
        self.footer.grid_rowconfigure(0, weight=1)
        self.footer.grid_columnconfigure(0, weight=1)

        # footer content
        self.last_activity_label = tk.Label(
            self.footer,
            text="last activity: xx:xx",
            fg=self.MUTED,
            bg=bg_color,
            font=self.FONT_SMALL,
            anchor="e",
        )
        self.last_activity_label.grid(row=0, column=1, padx=8, pady=16)


    def _apply_status_update(self, status: UIStatusText | None = None):

        # stops any ongoing animations
        self._stop_online_animation()
        self._stop_working_animation()
        self.status_dot.config(text="")

        # changes text
        if status == "online":
            self.rpa_status_label.config(text="online", fg=self.GREEN)
            self.jobs_normal_view.tkraise()
            self.status_dot.config(text="●")
            self._start_online_animation()

        elif status == "no network":
            self.rpa_status_label.config(text="no network", fg=self.RED)
            self.jobs_normal_view.tkraise()

        elif status == "working":
            self.rpa_status_label.config(text="working...", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()
            self._start_working_animation()

        elif status == "safestop":
            self.rpa_status_label.config(text="safestop", fg=self.RED)
            self.jobs_error_view.tkraise()

        elif status == "ooo":
            self.rpa_status_label.config(text="out-of-office", fg=self.YELLOW)
            self.jobs_normal_view.tkraise()


    def _apply_jobs_done_today(self, n) -> None:
        self.jobs_done_label.config(text=str(n))


    def _create_recording_overlay(self) -> None:
        #written by AI
        self.recording_win = tk.Toplevel(self.root)
        self.recording_win.withdraw()                # hidden at start
        self.recording_win.overrideredirect(True)    # no title/border
        self.recording_win.configure(bg="black")

        try:
            self.recording_win.attributes("-topmost", True)
        except Exception:
            pass

        width = self.RECORDING_WIDTH
        height = self.RECORDING_HEIGHT
        x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

        frame = tk.Frame(
            self.recording_win,
            bg="black",
            highlightbackground="#444444",
            highlightthickness=1,
            bd=0,
        )
        frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(
            frame,
            width=44,
            height=44,
            bg="black",
            highlightthickness=0,
            bd=0,
        )
        canvas.place(x=18, y=33)
        canvas.create_oval(4, 4, 40, 40, fill=self.RED, outline=self.RED)

        label = tk.Label(
            frame,
            text="RECORDING",
            fg="#FFFFFF",
            bg="black",
            font=self.FONT_RECORDING,
            anchor="w",
        )
        label.place(x=75, y=33)


    def _show_recording_overlay(self) -> None:
        #written by AI
        try:
            width = self.RECORDING_WIDTH
            height = self.RECORDING_HEIGHT
            x = self.root.winfo_screenwidth() - width - self.RECORDING_MARGIN_RIGHT
            y = (self.root.winfo_screenheight() // 2) - (height // 2)
            self.recording_win.geometry(f"{width}x{height}+{x}+{y}")

            self.recording_win.deiconify()
            self.recording_win.lift()

            try:
                self.recording_win.attributes("-topmost", True)
            except Exception:
                pass
        except Exception:
            pass


    def _hide_recording_overlay(self) -> None:
        # hides recording window
        try:
            self.recording_win.withdraw()
        except Exception:
            pass


    def _start_working_animation(self):
        if self._working_animation_after_id is None:
            self._animate_working()


    def _animate_working(self):
        #written by AI
        states = ["working", "working.", "working..", "working..."]
        self._working_dots = (self._working_dots + 1) % len(states)
        self.rpa_status_label.config(text=states[self._working_dots])
        self._working_animation_after_id = self.root.after(500, self._animate_working)


    def _stop_working_animation(self):
        if self._working_animation_after_id is not None:
            self.root.after_cancel(self._working_animation_after_id)
            self._working_animation_after_id = None
            self._working_dots = 0


    def _start_online_animation(self):
        if self._online_animation_after_id is None:
            self._online_pulse_index = 0
            self._animate_online()


    def _animate_online(self):
        # green pulse animation
        colors = [self.GREEN, self.GREEN_2, self.BG, self.GREEN_3, self.GREEN_2]
        color = colors[self._online_pulse_index]

        self.status_dot.config(fg=color)

        self._online_pulse_index = (self._online_pulse_index + 1) % len(colors)
        self._online_animation_after_id = self.root.after(1000, self._animate_online)


    def _stop_online_animation(self):
        if self._online_animation_after_id is not None:
            self.root.after_cancel(self._online_animation_after_id)
            self._online_animation_after_id = None


    def _append_ui_log(self, log_line: str, blank_line_before: bool = False) -> None:

        self.log_text.config(state="normal")  # open for edit
        now = datetime.datetime.now().strftime("%H:%M")

        if blank_line_before:
            self.log_text.insert("end", "\n")

        self.log_text.insert("end", f"[{now}] {log_line}\n")

        self.log_text.config(state="disabled")  # closing edit
        self.log_text.see("end")

    # all 'post_...' below are thread-safe wrappers
    def post_status_update(self, status: UIStatusText) -> None:
        self.root.after(0, lambda: self._apply_status_update(status))

    def post_log_line(self, text: str, blank_line_before: bool = False) -> None:
        self.root.after(0, lambda: self._append_ui_log(text, blank_line_before))

    def post_show_recording_overlay(self) -> None:
        self.root.after(0, self._show_recording_overlay)

    def post_hide_recording_overlay(self) -> None:
        self.root.after(0, self._hide_recording_overlay)

    def post_jobs_done_today(self, n: int) -> None:
        self.root.after(0, lambda: self._apply_jobs_done_today(n))

    def post_shutdown(self) -> None:
        self.root.after(0, self.shutdown)


# ============================================================
# MAIN ENTRYPOINT
# ============================================================

class RobotRuntime:
    """Main orchestration runtime."""

    WATCHDOG_TIMEOUT = 10  # demo-friendly watchdog timeout (seconds). Max wait time for RPA tool
    POLL_INTERVAL = 1   # demo-friendly poll interval for runtime_loop() 
    QUERYFLOW_POLLINTERVAL = 1  # demo-friendly poll interval for queries (seconds)

    def __init__(self, ui):

        self.prev_ui_status = None
        self.next_queryflow_check_time = 0
        self.prev_ipc_state: IpcState | None = None
        self.watchdog_started_at: float | None = None

        self.ui = ui
        self.logger = LoggerService(self.ui)

        self.handover_repo = HandoverRepository(self.logger)
        self.friends_repo = FriendsRepository()
        self.audit_repo = AuditRepository(self.logger)
        self.network_service = NetworkService(self.logger)
        self.recording_service = RecordingService(self.logger)
        
        self.mail_backend_personal = ExampleMailBackend(self.logger, "personal_inbox")
        self.mail_backend_shared = ExampleMailBackend(self.logger, "shared_inbox")
        self.erp_backend = ExampleErpBackend()

        self.job_handlers = {
            "ping": ExamplePingJobHandler(self.logger),
            "job1": ExampleJob1Handler(self.logger), 
            "job2": ExampleJob2Handler(self.logger), 
            "job3": ExampleJob3Handler(self.logger, self.erp_backend),
            }
        
        self.notification_service = UserNotificationService(self.mail_backend_personal, self.recording_service.RECORDINGS_DESTINATION_FOLDER, self.WATCHDOG_TIMEOUT)
        self.pre_handover_executor = PreHandoverExecutor(self.logger, self.update_ui_status, self.ui.post_show_recording_overlay, self.generate_job_id, self.recording_service, self.audit_repo, self.notification_service, self.mail_backend_personal, self.mail_backend_shared,)
        self.mail_flow = MailFlow(self.logger, self.friends_repo, self.is_within_operating_hours, self.network_service, self.job_handlers, self.pre_handover_executor, self.mail_backend_personal, self.mail_backend_shared)
        self.query_flow = QueryFlow(self.logger, self.audit_repo, self.job_handlers, self.pre_handover_executor, self.is_within_operating_hours, self.erp_backend)
        self.post_handover_finalizer = PostHandoverFinalizer(self.logger, self.audit_repo, self.job_handlers, self.recording_service, self.ui.post_hide_recording_overlay, self.mail_backend_personal, self.mail_backend_shared, self.notification_service)
        self.safestop_controller = SafeStopController(self.logger, self.recording_service, self.ui.post_hide_recording_overlay, self.ui.post_status_update, self.ui.post_shutdown, self.mail_backend_personal, self.audit_repo, self.generate_job_id, self.friends_repo, self.notification_service, self.poll_for_rpatool_stop_flag,) 

        
    def initialize_runtime(self,):
        self.logger.system(f"RuntimeThread started, version={VERSION}, pid={os.getpid()}")

        # write 'idle' to allow for manual start of main.py (not the intended way)
        active_job=ActiveJob(ipc_state="idle")
        self.handover_repo.write(active_job)

        active_job = self.handover_repo.read()
        if active_job.ipc_state != "idle":
            raise RuntimeError(f"Expected handover.json to start in idle, got {active_job.ipc_state}")
        
        # cleanup
        for fn in ["stop.flag", "restart.flag"]:
            try: os.remove(fn)
            except Exception: pass

        atexit.register(self.recording_service.stop) # during normal exit
 
        self.network_service.has_network_access()
        self.recording_service.stop() # stop any active recordings since last session
        self.recording_service.cleanup_aborted_recordings()
        self.friends_repo.reload_if_modified()
        self.audit_repo.ensure_db_exists()
        self.refresh_jobs_done_counter()

        # Retry missing final replies from previous crash/restart.
        if self.audit_repo.get_pending_reply_jobs():
            self.safestop_controller.send_recovery_replies(from_initialize=True)


    def runtime_loop(self) -> None:
        active_job: ActiveJob | None = None
        
        try:          
            self.initialize_runtime()

            while True:
                self.poll_for_rpatool_stop_flag()

                active_job = self.handover_repo.read()
                ipc_state = active_job.ipc_state
                job_id = active_job.job_id
                
                # dispatch
                if ipc_state == "idle":             # Runtime owns the workflow
                    self.poll_job_intake()

                elif ipc_state == "job_queued":     # RPA tool owns the workflow
                    pass

                elif ipc_state == "job_running":    # RPA tool owns the workflow
                    pass

                elif ipc_state == "job_verifying":  # Runtime owns the workflow
                    self.finalize_current_job(active_job, job_id)

                elif ipc_state == "safestop":       # Runtime owns the workflow
                    raise RuntimeError(f"safestop from RPA tool (RPA_TOOL_CRASH)")   

                self._handle_state_transition(ipc_state, job_id)
                self._enforce_watchdog(ipc_state)

                time.sleep(self.POLL_INTERVAL)


        except Exception as err_short:  # policy to safe-stop on errors
            err_with_traceback = traceback.format_exc()

            try: active_job = self.handover_repo.read()
            except Exception: active_job = None

            self.safestop_controller.run_degraded_mode(
                str(err_short),
                err_with_traceback,
                active_job,)


    def refresh_jobs_done_counter(self):

        count = self.audit_repo.count_done_jobs_today()
        self.ui.post_jobs_done_today(count)


    def _handle_state_transition(self, ipc_state, job_id) -> None:
        if ipc_state != self.prev_ipc_state:
            transition_message=f"state transition detected by CPU-poll: {self.prev_ipc_state} -> {ipc_state}"

            #if not self.handover_repo.is_valid_ipc_transition(self.prev_ipc_state, ipc_state):
            #    raise RuntimeError(f"invalid {transition_message}")

            self.update_ui_status(ipc_state)
            self.logger.system(transition_message, job_id)

            if ipc_state == "job_running":
                self.audit_repo.update_job(job_id=job_id, job_status="RUNNING")

            # note handover time or last RPA tool state transition
            if ipc_state in ("job_queued", "job_running"):
                self.watchdog_started_at =  time.time()
            else:
                self.watchdog_started_at = None
        

    def _enforce_watchdog(self, ipc_state):
        self.prev_ipc_state = ipc_state

        if not self.watchdog_started_at:
            return

        if time.time() - self.watchdog_started_at <= self.WATCHDOG_TIMEOUT:
            return

        if ipc_state == "job_queued":
            raise RuntimeError(f"nothing started within {self.WATCHDOG_TIMEOUT} seconds (PRE_HANDOVER_CRASH)")

        if ipc_state == "job_running":
            raise RuntimeError(f"not completed within {self.WATCHDOG_TIMEOUT} seconds (RPA_TOOL_CRASH)")    


    def update_ui_status(self, ipc_state=None, forced_status=None) -> None:
               
        if forced_status is not None:
            if forced_status not in get_args(UIStatusText):
                raise ValueError(f"unknown forced_status: {forced_status}")
            ui_status: UIStatusText = forced_status

        else:
            if ipc_state is not None and ipc_state not in get_args(IpcState):
                raise ValueError(f"unknown ipc_state: {ipc_state}")

            if ipc_state == "safestop":
                ui_status = "safestop"

            elif ipc_state in ("job_queued", "job_running", "job_verifying"):
                ui_status = "working"

            elif self.network_service.network_state is False:
                ui_status = "no network"

            elif not self.is_within_operating_hours():
                ui_status = "ooo"

            else:
                ui_status = "online"

        if self.prev_ui_status != ui_status:
            self.ui.post_status_update(ui_status)
            self.prev_ui_status = ui_status


    def poll_job_intake(self) -> bool:
        ''' job intake logic '''
        try:
            
            # 1. Mail first (priority)
            poll_result = self.mail_flow.poll_once()
            if poll_result.handled_anything:
                if poll_result.active_job:
                    self.handover_repo.write(poll_result.active_job)
                return True
            
        
            # 2. Query (or other scheduled) jobs
            now = time.time()
            if now < self.next_queryflow_check_time:
                return False

            poll_result = self.query_flow.poll_once()

            if poll_result.active_job:
                self.handover_repo.write(poll_result.active_job)
                return True

            # prolong intervall if no new match
            self.next_queryflow_check_time = now + self.QUERYFLOW_POLLINTERVAL 
            return False


        except Exception as e:
            raise RuntimeError(f"PRE_HANDOVER_CRASH: {e}")

        
    def generate_job_id(self) -> int:
        ''' unique id for all jobs '''

        job_id = int(datetime.datetime.now().strftime("%Y%m%d%H%M%S"))

        last_job_id = self.audit_repo.get_latest_job_id()
        job_id = max(job_id, last_job_id + 1)

        self.logger.system(f"assigned job_id", job_id)
        return job_id

    
    def is_within_operating_hours(self) -> bool:

        now = datetime.datetime.now().time()
        result = datetime.time(5,0) <= now <= datetime.time(23,0) # eg. operating hours 05:00 to 23:00
        
        return result
        

    def finalize_current_job(self, active_job: ActiveJob, job_id) -> None:

        self.logger.system(f"finalizing {active_job.job_type} with payload {active_job.rpatool_payload}", job_id) # only store safe data in log
        
        self.post_handover_finalizer.poll_once(active_job, job_id)

        self.refresh_jobs_done_counter()

        self.handover_repo.write(ActiveJob(
            ipc_state="idle",
        ))


    def poll_for_rpatool_stop_flag(self,):
        ''' to stop main.py on operator manual stop on RPA tool '''

        stopflag = "stop.flag"
 
        if os.path.isfile(stopflag):
            try: os.remove(stopflag)
            except Exception: pass

            self.logger.system(f"found {stopflag}")
            
            try: self.ui.post_shutdown() #request soft-exit
            except Exception: os._exit(1)
            
            time.sleep(3)
            os._exit(0)  #kill if still alive after 3 sec 


    def request_shutdown(self) -> None:
        try:
            self.recording_service.stop()
        except Exception:
            pass


def main() -> None:
    ''' run Dashboard UI in main thread and 'the rest' async '''
    ui = DashboardUI()
    robot_runtime = RobotRuntime(ui)

    ui.set_shutdown_callback(robot_runtime.request_shutdown)

    threading.Thread(target=robot_runtime.runtime_loop, daemon=True).start() # 'the rest'

    ui.run()



if __name__ == "__main__":
    main()