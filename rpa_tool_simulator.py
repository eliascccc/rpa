from __future__ import annotations

import datetime
import json
import os
import platform
import subprocess
import sys
import threading
import time
from pathlib import Path

from openpyxl import load_workbook  # type: ignore


class RPAToolSimulator:
    """
    Simulates the external RPA tool.

    Mental model:
    * Start this script = open the RPA tool
    * Press 1 = press RUN in the RPA tool
    * Press 2 = press STOP in the RPA tool
    """

    def __init__(self) -> None:
        self.last_command: str | None = None

    def run(self):
        self.log_system("RPAToolSimulator alive")
        print("Button 1 and 2 represents the START/RUN and the STOP button in the RPA Tool.")

        # input runs in separate thread
        threading.Thread(target=self._command_loop, daemon=True).start()

        while True:
            print("[stopped] Press 1 to start the robot.")
            self.wait_for_command("1")
            self.run_robot()

    def run_robot(self):
        # --------------------------------------------------
        # Cold start policy: reset handover.json on startup
        # --------------------------------------------------
        handover_data = {"ipc_state": "idle"}
        with open("handover.json", "w", encoding="utf-8") as f:
            json.dump(handover_data, f, indent=2)

        # --------------------------------------------------
        # Start main.py (RobotRuntime) async
        # --------------------------------------------------
        self.start_runtime_in_new_terminal()
        print("[started] Press 2 to request stop.")

        # --------------------------------------------------
        #  Enter normal polling loop
        # --------------------------------------------------
        while True:
            # operator stop request
            if self.last_command == "2":
                self.last_command = None
                Path("stop.flag").write_text("", encoding="utf-8")
                self.log_system("stop.flag written by RPAToolSimulator")
                return

            time.sleep(1)

            try:
                # read handover
                with open("handover.json", "r", encoding="utf-8") as f:
                    handover_data = json.load(f)

                ipc_state = handover_data.get("ipc_state")
                if ipc_state != "job_queued":
                    continue

               # claim workflow if "job_queued"
                handover_data["ipc_state"] = "job_running"
                with open("handover.json", "w", encoding="utf-8") as f:
                    json.dump(handover_data, f, indent=2)

                # identify job
                job_type = handover_data.get("job_type")
                job_id = handover_data.get("job_id")
                rpatool_payload = handover_data.get("rpatool_payload")

                if rpatool_payload is None:
                    raise ValueError("did your forgot something?")

                time.sleep(2)  # simulate processing time

                # JOB1
                if job_type == "job1":
                    # retrive job-specific data 
                    erp_order_number = rpatool_payload.get("order_number")
                    new_qty = rpatool_payload.get("target_order_qty")

                    # simulation of job1 screenactiviy
                    self.log_system("activities on screen_1 in ERP completed", job_id)
                    self.log_system("activities on screen_2 in ERP completed", job_id)
                    self.simulate_rpa_result_job1(erp_order_number, new_qty)

                    new_ipc_state = "job_verifying"
                    

                # JOB3
                elif job_type == "job3":
                    erp_order_number = rpatool_payload.get("source_ref")
                    new_qty = rpatool_payload.get("target_order_qty")

                    self.log_system("activities on screen_1 in ERP completed", job_id)
                    self.log_system("activities on screen_2 in ERP completed", job_id)
                    self.simulate_rpa_result_job1(erp_order_number, new_qty) # use job1 example         

                    new_ipc_state = "job_verifying"

                # PING
                elif job_type == "ping":
                    if platform.system() == "Windows":
                        import winsound
                        winsound.Beep(1000, 300)  # type: ignore
                    elif platform.system() == "Linux":
                        print("\a", end="", flush=True)

                    self.log_system("made a ping", job_id)
                    new_ipc_state = "job_verifying"

                # UNKNOWN JOB
                else:
                    self.log_system(f"no logic for job_type={job_type}", job_id)
                    new_ipc_state = "safestop"

                # handover back to RobotRuntime
                handover_data["ipc_state"] = new_ipc_state
                with open("handover.json", "w", encoding="utf-8") as f:
                    json.dump(handover_data, f, indent=2)

                self.log_system(
                    f"RPAToolSimulator done, ipc_state job_running -> {new_ipc_state}",
                    job_id,
                )

            except Exception as e:
                self.log_system(f"crash in polling loop: {e}")

                try:
                    handover_data["ipc_state"] = "safestop"
                except Exception:
                    handover_data = {"ipc_state": "safestop"}

                with open("handover.json", "w", encoding="utf-8") as f:
                    json.dump(handover_data, f, indent=2)

                print("RPA tool entered safestop.")
                return

    def wait_for_command(self, expected: str):
        while self.last_command != expected:
            time.sleep(0.1)
        self.last_command = None

    def _command_loop(self):
        while True:
            try:
                cmd = input("> ").strip().lower()
                if cmd in ("1", "2"):
                    self.last_command = cmd
                else:
                    print("Unknown command. Use 1 or 2.")
            except Exception as e:
                print(f"Command loop error: {e}")

    def simulate_rpa_result_job1(self, erp_order_number: str, new_qty: int, path="Example_ERP_table.xlsx"):
        assert erp_order_number is not None
        assert new_qty is not None

        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == str(erp_order_number):
                row[1].value = int(new_qty)  # type: ignore
                wb.save(path)
                wb.close()
                return True

        wb.close()
        return False

    def start_runtime_in_new_terminal(self):
        python_exe = sys.executable
        script_path = os.path.abspath("main.py")

        if platform.system() == "Windows":
            subprocess.Popen([
                "powershell",
                "-NoExit",
                "-Command",
                f'& "{python_exe}" "{script_path}"'
            ])
            return

        python_cmd = f'"{python_exe}" "{script_path}"'
        terminal_candidates = [
            ["gnome-terminal", "--", "bash", "-lc", f"{python_cmd}; exec bash"],
            ["xfce4-terminal", "--hold", "-e", python_cmd],
            ["konsole", "-e", "bash", "-lc", f"{python_cmd}; exec bash"],
            ["xterm", "-hold", "-e", python_cmd],
        ]

        for cmd in terminal_candidates:
            try:
                subprocess.Popen(cmd)
                return
            except FileNotFoundError:
                continue

        raise RuntimeError("No supported terminal emulator found")

    def log_system(self, event_text: str, job_id=None):
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        job_id_text = "" if job_id is None else str(job_id)

        log_line = f"{timestamp} | rpa | job_id={job_id_text} | RPAToolSimulator...() | {event_text}"

        with open("system.log", "a", encoding="utf-8") as f:
            f.write(log_line + "\n")
            f.flush()


def main():
    if not os.path.isfile("main.py"):
        raise RuntimeError("Place this file in main.py directory")

    RPAToolSimulator().run()


if __name__ == "__main__":
    main()